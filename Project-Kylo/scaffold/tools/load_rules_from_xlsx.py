from __future__ import annotations

import json
import os
import sys
import uuid
from dataclasses import dataclass
from hashlib import sha256
from typing import Any, Dict, List, Optional, Tuple

import psycopg2
from openpyxl import load_workbook


def _strip(val: Any) -> str:
    return ("" if val is None else str(val)).strip()


def _title_to_company_id(title: str) -> str:
    # Expect "Pending Rules – <CID>"
    parts = title.split("–", 1)
    return parts[1].strip() if len(parts) == 2 else title.strip()


def _normalize_header(name: str) -> str:
    # Lowercase, replace spaces/dashes with underscores, collapse multiple underscores
    n = (name or "").strip().lower()
    for ch in [" ", "-", "\u2013", "\u2014"]:
        n = n.replace(ch, "_")
    while "__" in n:
        n = n.replace("__", "_")
    return n


def _canonicalize_headers(headers: List[str]) -> Dict[str, str]:
    """Map observed headers to canonical names expected by the loader.

    Accepts common variants like "Unique Source" -> "Source",
    "Target sheet" -> "Target_Sheet", "Target Header" -> "Target_Header".
    Returns a dict mapping CanonicalName -> ObservedHeader.
    """
    alias_to_canonical: Dict[str, str] = {
        # Canonical required/known columns
        "approved": "Approved",
        "source": "Source",
        "unique_source": "Source",
        "company_id": "Company_ID",
        "date": "Date",
        "amount": "Amount",
        "target_sheet": "Target_Sheet",
        "target_sheet_": "Target_Sheet",
        "target__sheet": "Target_Sheet",
        "target_header": "Target_Header",
        "match_notes": "Match_Notes",
        "rule_id": "Rule_ID",
        "processed_at": "Processed_At",
        "row_hash": "Row_Hash",
    }

    observed_norm_to_raw: Dict[str, str] = { _normalize_header(h): h for h in headers }
    mapping: Dict[str, str] = {}
    for norm, raw in observed_norm_to_raw.items():
        if norm in alias_to_canonical:
            mapping[ alias_to_canonical[norm] ] = raw
        # Special-case: "target sheet" often appears as two words
        if norm == "target_sheet" and "Target_Sheet" not in mapping:
            mapping["Target_Sheet"] = raw
        if norm == "target_header" and "Target_Header" not in mapping:
            mapping["Target_Header"] = raw
        if norm == "unique_source" and "Source" not in mapping:
            mapping["Source"] = raw
    return mapping


def _row_to_rule_json(header: List[str], row: List[Any], fallback_company: str) -> Tuple[str, Dict[str, Any]]:
    canonical = _canonicalize_headers(header)
    def get_val(canonical_name: str) -> Any:
        raw = canonical.get(canonical_name)
        if raw is None:
            return None
        try:
            idx = header.index(raw)
        except ValueError:
            return None
        return row[idx] if idx < len(row) else None

    rule_id = _strip(get_val("Rule_ID")) or str(uuid.uuid4())
    rule_json: Dict[str, Any] = {
        "date": _strip(get_val("Date")),
        "source": _strip(get_val("Source")).lower(),
        "company_id": _strip(get_val("Company_ID")) or fallback_company,
        "amount": _strip(get_val("Amount")),
        "target_sheet": _strip(get_val("Target_Sheet")),
        "target_header": _strip(get_val("Target_Header")),
        "match_notes": _strip(get_val("Match_Notes")),
    }
    return rule_id, rule_json


def _compute_rule_hash(rule_json: Dict[str, Any]) -> str:
    payload = json.dumps(rule_json, sort_keys=True, separators=(",", ":"))
    return sha256(payload.encode("utf-8")).hexdigest()


@dataclass
class CompanyLoad:
    company_id: str
    rules: List[Tuple[str, Dict[str, Any]]]  # (rule_id, rule_json)


def parse_workbook(path: str) -> List[CompanyLoad]:
    wb = load_workbook(path, read_only=True, data_only=True)
    loads: Dict[str, CompanyLoad] = {}
    for ws in wb.worksheets:
        title = ws.title
        company_id = _title_to_company_id(title)
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        canon = _canonicalize_headers(header)
        # Validate minimal columns: Source + Target_Sheet + Target_Header must be present (via aliases)
        if not ("Source" in canon and "Target_Sheet" in canon and "Target_Header" in canon):
            continue
        col_idx = {name: i for i, name in enumerate(header)}
        company = loads.setdefault(company_id, CompanyLoad(company_id=company_id, rules=[]))
        for r in rows[1:]:
            if not r:
                continue
            # Approved column is optional; default TRUE when missing
            if "Approved" in canon:
                idx = col_idx.get(canon["Approved"], None)
                approved_val = r[idx] if (idx is not None and idx < len(r)) else None
                approved = str(approved_val).strip().upper() == "TRUE"
            else:
                approved = True
            if not approved:
                continue
            rule_id, rule_json = _row_to_rule_json(header, list(r), company_id)
            company.rules.append((rule_id, rule_json))
    return list(loads.values())


def load_into_db(dsn: str, company_load: CompanyLoad) -> Tuple[int, int]:
    inserted = 0
    updated = 0
    with psycopg2.connect(dsn) as conn:
        conn.autocommit = False
        with conn.cursor() as cur:
            for rule_id, rule_json in company_load.rules:
                rule_hash = _compute_rule_hash(rule_json)
                cur.execute(
                    """
                    INSERT INTO app.rules_active (rule_id, rule_json, applied_at, rule_hash)
                    VALUES (%s, %s::jsonb, now(), %s)
                    ON CONFLICT (rule_id) DO UPDATE
                    SET rule_json = EXCLUDED.rule_json,
                        applied_at = now(),
                        rule_hash = EXCLUDED.rule_hash
                    RETURNING (xmax = 0)::int
                    """,
                    (rule_id, json.dumps(rule_json), rule_hash),
                )
                was_insert = cur.fetchone()[0] == 1
                if was_insert:
                    inserted += 1
                else:
                    updated += 1
        conn.commit()
    return inserted, updated


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python -m scaffold.tools.load_rules_from_xlsx <Rules.xlsx>", file=sys.stderr)
        return 2
    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}", file=sys.stderr)
        return 2

    dsn_map_env = os.environ.get("KYLO_DB_DSN_MAP")
    if not dsn_map_env:
        print("KYLO_DB_DSN_MAP not set; expected JSON like {\"company_a\":\"postgresql://...\"}", file=sys.stderr)
        return 2
    try:
        dsn_map: Dict[str, str] = json.loads(dsn_map_env)
    except Exception as e:
        print(f"Invalid KYLO_DB_DSN_MAP: {e}", file=sys.stderr)
        return 2

    loads = parse_workbook(xlsx_path)
    total_ins = 0
    total_upd = 0
    for cl in loads:
        dsn = dsn_map.get(cl.company_id)
        if not dsn:
            print(f"Skip company {cl.company_id}: no DSN in KYLO_DB_DSN_MAP", file=sys.stderr)
            continue
        ins, upd = load_into_db(dsn, cl)
        print(f"{cl.company_id}: inserted={ins}, updated={upd}")
        total_ins += ins
        total_upd += upd
    print(f"Done. total_inserted={total_ins}, total_updated={total_upd}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


