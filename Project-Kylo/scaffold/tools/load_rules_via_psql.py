from __future__ import annotations

import argparse
import json
import os
import uuid
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from .seed_pending_from_xlsx import (
    _canonicalize_headers,
    _build_company_alias_map,
    _norm_key,
    _strip,
)
from .sheets_cli import load_companies_config


def sanitize_db_name_fragment(s: str) -> str:
    return "".join(ch for ch in s.lower() if ch.isalnum() or ch in ("_",))


def company_to_db_name(company: Dict[str, Any]) -> str:
    # Prefer rules_namespace (e.g., "puffin"), fallback to sanitized first token of company_id
    ns = (company.get("rules_namespace") or "").strip()
    if ns:
        return f"kylo_{sanitize_db_name_fragment(ns)}"
    cid = (company.get("company_id") or "").strip()
    base = cid.split(" ")[0] if cid else cid
    return f"kylo_{sanitize_db_name_fragment(base)}"


def parse_active_rows(xlsx_path: str, companies: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    out: Dict[str, List[Dict[str, Any]]] = {}
    alias_map = _build_company_alias_map(companies)
    for ws in wb.worksheets:
        title_company = ws.title.split("–", 1)[-1].strip() if "–" in ws.title else ws.title.strip()
        company_id = alias_map.get(_norm_key(title_company), title_company)

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_raw = [str(c).strip() if c is not None else "" for c in rows[0]]
        canon = _canonicalize_headers(header_raw)
        if not ("Source" in canon and "Target_Sheet" in canon and "Target_Header" in canon):
            continue

        def get(cname: str, row: Tuple[Any, ...]) -> Any:
            raw = canon.get(cname)
            if raw is None:
                return None
            try:
                idx = header_raw.index(raw)
            except ValueError:
                return None
            return row[idx] if idx < len(row) else None

        active_rows: List[Dict[str, Any]] = []
        consecutive_blank = 0
        for r in rows[1:]:
            if not r:
                consecutive_blank += 1
                if consecutive_blank > 50:
                    break
                continue
            consecutive_blank = 0

            src = _strip(get("Source", r)).lower()
            tgt_sheet = _strip(get("Target_Sheet", r))
            tgt_header = _strip(get("Target_Header", r))
            if not (src and tgt_sheet and tgt_header):
                continue
            # Minimal active rule JSON; treat as approved when loading actives
            row_json: Dict[str, Any] = {
                "source": src,
                "target_sheet": tgt_sheet,
                "target_header": tgt_header,
                "approved": True,
            }
            active_rows.append(row_json)
        if active_rows:
            out.setdefault(company_id, []).extend(active_rows)
    return out


def json_sql_literal(obj: Any) -> str:
    s = json.dumps(obj, separators=(",", ":"))
    return s.replace("'", "''")


def build_sql_for_company(rows: List[Dict[str, Any]]) -> str:
    # One INSERT per row; rule_hash computed in DB from JSON text; rule_id supplied from Python uuid4
    stmts: List[str] = []
    for row in rows:
        row_json_sql = json_sql_literal(row)
        rid = str(uuid.uuid4())
        stmts.append(
            (
                "INSERT INTO app.rules_active (rule_id, rule_json, applied_at, rule_hash) "
                "VALUES ('{rid}', '{j}'::jsonb, now(), "
                "encode(digest('{j}', 'sha256'), 'hex')::char(64)) "
                "ON CONFLICT (rule_id) DO UPDATE SET "
                "rule_json = EXCLUDED.rule_json, applied_at = now(), rule_hash = EXCLUDED.rule_hash;"
            ).format(rid=rid, j=row_json_sql)
        )
    return "\n".join(stmts) + "\n"


def main() -> None:
    p = argparse.ArgumentParser(description="Load Active rules into company DBs via docker exec psql")
    p.add_argument("--xlsx", default="Rules.xlsx")
    p.add_argument("--config", default=os.path.join("config", "companies.json"))
    p.add_argument("--container", default="kylo-pg")
    p.add_argument("--pguser", default="postgres")
    p.add_argument("--pgpass", default="kylo")
    args = p.parse_args()

    cfg = load_companies_config(args.config)
    companies: List[Dict[str, Any]] = cfg.get("companies", [])
    if not companies:
        raise SystemExit("No companies in config")

    by_company = parse_active_rows(args.xlsx, companies)

    tmp_dir = os.path.join("db", "tmp_rules")
    os.makedirs(tmp_dir, exist_ok=True)

    for company in companies:
        cid = company.get("company_id")
        dbname = company_to_db_name(company)
        rows = by_company.get(cid, [])
        print(json.dumps({"company": cid, "db": dbname, "active_rows": len(rows)}))
        if not rows:
            continue
        sql = build_sql_for_company(rows)
        local_path = os.path.join(tmp_dir, f"{dbname}.sql")
        with open(local_path, "w", encoding="utf-8") as f:
            f.write(sql)
        # Copy and execute
        container_path = f"/tmp/{dbname}.sql"
        os.system(f"docker cp {local_path} {args.container}:{container_path}")
        os.system(
            f"docker exec -e PGPASSWORD={args.pgpass} {args.container} psql -U {args.pguser} -d {dbname} -v ON_ERROR_STOP=1 -f {container_path}"
        )


if __name__ == "__main__":
    main()


