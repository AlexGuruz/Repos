from __future__ import annotations

import argparse
import json
import os
from typing import Any, Dict, List

from openpyxl import load_workbook
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build


def _get_service(sa_path: str):
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    from google.oauth2.service_account import Credentials
    return build("sheets", "v4", credentials=Credentials.from_service_account_file(sa_path, scopes=scopes))


def extract_spreadsheet_id(url_or_id: str) -> str:
    if "/d/" in url_or_id:
        return url_or_id.split("/d/")[1].split("/")[0]
    return url_or_id


def _to_cell(v: Any) -> Dict[str, Any]:
    return {"userEnteredValue": {"stringValue": "" if v is None else str(v)}}


def _to_row(values: List[Any]) -> Dict[str, Any]:
    return {"values": [_to_cell(v) for v in values]}


def main() -> int:
    ap = argparse.ArgumentParser(description="Seed consolidated 'CID Pending' tabs from Rules.xlsx")
    ap.add_argument("--xlsx", default="Rules.xlsx")
    ap.add_argument("--spreadsheet-id", required=True, help="Consolidated spreadsheet (title, id, or URL)")
    ap.add_argument("--service-account", default=os.path.join("secrets", "service_account.json"))
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"File not found: {args.xlsx}")
        return 2
    if not os.path.exists(args.service_account):
        print(f"Service account not found: {args.service_account}")
        return 2

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    sid = extract_spreadsheet_id(args.spreadsheet_id)
    service = _get_service(args.service_account)

    # Fetch titles→ids map
    meta = service.spreadsheets().get(spreadsheetId=sid).execute()
    titles_to_ids = {s.get("properties", {}).get("title"): int(s.get("properties", {}).get("sheetId")) for s in meta.get("sheets", [])}

    wrote = 0
    for ws in wb.worksheets:
        title = ws.title.strip()
        if not title:
            continue
        pending_title = f"{title} Pending"
        sheet_id = titles_to_ids.get(pending_title)
        if sheet_id is None:
            # skip sheets without a matching consolidated Pending tab
            continue
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        body = rows[1:]
        # Write to consolidated Pending: clear body then write
        body_rows = [[str(c) if c is not None else "" for c in r[:7]] for r in body]
        reqs = [
            {"updateCells": {"range": {"sheetId": sheet_id, "startRowIndex": 1}, "fields": "userEnteredValue"}},
            {
                "updateCells": {
                    "range": {"sheetId": sheet_id, "startRowIndex": 1, "startColumnIndex": 0},
                    "fields": "userEnteredValue",
                    "rows": [_to_row(r) for r in body_rows],
                }
            },
        ]
        service.spreadsheets().batchUpdate(spreadsheetId=sid, body={"requests": reqs}).execute()
        wrote += len(body_rows)
    print(json.dumps({"spreadsheetId": sid, "wrote_rows": wrote}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


