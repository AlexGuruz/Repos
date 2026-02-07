from __future__ import annotations

import argparse
import json
import os
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from .load_rules_from_xlsx import (
    _compute_rule_hash,  # reuse hashing from loader
    _title_to_company_id,
)
from .sheets_batch import (
    _get_service,
    resolve_sheet_ids,
    ensure_setup_batch,
)
from .sheets_cli import load_companies_config, extract_spreadsheet_id_from_url


PENDING_HEADERS: List[str] = [
    "Row_Hash",
    "Date",
    "Source",
    "Company_ID",
    "Amount",
    "Target_Sheet",
    "Target_Header",
    "Match_Notes",
    "Approved",
    "Rule_ID",
    "Processed_At",
]


def _strip(val: Any) -> str:
    return ("" if val is None else str(val)).strip()


def _normalize_header(name: str) -> str:
    n = (name or "").strip().lower()
    for ch in [" ", "-", "\u2013", "\u2014"]:
        n = n.replace(ch, "_")
    while "__" in n:
        n = n.replace("__", "_")
    return n


def _canonicalize_headers(headers: List[str]) -> Dict[str, str]:
    alias_to_canonical: Dict[str, str] = {
        "approved": "Approved",
        "source": "Source",
        "unique_source": "Source",
        "company_id": "Company_ID",
        "date": "Date",
        "amount": "Amount",
        "target_sheet": "Target_Sheet",
        "target_header": "Target_Header",
        "match_notes": "Match_Notes",
        "rule_id": "Rule_ID",
    }
    observed_norm_to_raw: Dict[str, str] = {_normalize_header(h): h for h in headers}
    mapping: Dict[str, str] = {}
    for norm, raw in observed_norm_to_raw.items():
        if norm in alias_to_canonical:
            mapping[alias_to_canonical[norm]] = raw
    # Convenience fallbacks
    if "Target_Sheet" not in mapping and "target sheet" in (h.lower() for h in headers):
        for h in headers:
            if h.lower() == "target sheet":
                mapping["Target_Sheet"] = h
                break
    if "Target_Header" not in mapping and "target header" in (h.lower() for h in headers):
        for h in headers:
            if h.lower() == "target header":
                mapping["Target_Header"] = h
                break
    if "Source" not in mapping and "unique source" in (h.lower() for h in headers):
        for h in headers:
            if h.lower() == "unique source":
                mapping["Source"] = h
                break
    return mapping


def _norm_key(s: str) -> str:
    return "".join(ch for ch in (s or "").lower() if ch.isalnum())


def _build_company_alias_map(companies: List[Dict[str, Any]]) -> Dict[str, str]:
    """Return mapping of normalized alias -> canonical company_id."""
    alias_map: Dict[str, str] = {}
    for c in companies:
        cid = c.get("company_id", "")
        ns = c.get("rules_namespace", "")
        aliases = {cid, ns, (cid.split(" ")[0] if cid else "")}
        for a in aliases:
            k = _norm_key(a)
            if k:
                alias_map[k] = cid
    return alias_map


def parse_rules_xlsx(xlsx_path: str, companies: List[Dict[str, Any]]) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """Return dict company_id -> {"pending": [...], "active": [...]}.

    pending: rows with Source present but missing Target_Sheet or Target_Header
    active: rows with all of Source, Target_Sheet, Target_Header
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    out: Dict[str, Dict[str, List[Dict[str, Any]]]] = {}
    alias_map = _build_company_alias_map(companies)
    for ws in wb.worksheets:
        title_company = _title_to_company_id(ws.title)
        # Map sheet title to canonical company_id using aliases (e.g., "PUFFIN" -> "PUFFIN PURE")
        company_id = alias_map.get(_norm_key(title_company), title_company)
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_raw = [str(c).strip() if c is not None else "" for c in rows[0]]
        canon = _canonicalize_headers(header_raw)
        if not ("Source" in canon and "Target_Sheet" in canon and "Target_Header" in canon):
            # Not a rules sheet; skip
            continue
        def get(cname: str, row: Tuple[Any, ...]) -> Any:
            raw = canon.get(cname)
            if raw is None:
                return None
            try:
                idx = header_raw.index(raw)
            except ValueError:
                return None
            return row[idx] if idx < len(row) else None

        pending_rows: List[Dict[str, Any]] = []
        active_rows: List[Dict[str, Any]] = []
        consecutive_blank = 0
        for r in rows[1:]:
            if not r:
                consecutive_blank += 1
                if consecutive_blank > 50:
                    break
                continue
            consecutive_blank = 0
            # Build a minimal rule_json to compute Row_Hash
            src = _strip(get("Source", r)).lower()
            tgt_sheet = _strip(get("Target_Sheet", r))
            tgt_header = _strip(get("Target_Header", r))
            # If Source is blank, skip entirely
            if not src:
                continue

            approved_raw = get("Approved", r)
            if approved_raw is not None and str(approved_raw).strip().upper() not in ("TRUE", "1", "Y", "YES"):
                # Respect explicit unapproved rows
                continue

            rule_json = {
                "date": _strip(get("Date", r)),
                "source": src,
                "company_id": company_id,
                "amount": _strip(get("Amount", r)),
                "target_sheet": tgt_sheet,
                "target_header": tgt_header,
                "match_notes": _strip(get("Match_Notes", r)),
            }
            row_hash = _compute_rule_hash(rule_json)
            row_obj = {
                "Row_Hash": row_hash,
                "Date": rule_json["date"],
                "Source": rule_json["source"],
                "Company_ID": company_id,
                "Amount": rule_json["amount"],
                "Target_Sheet": rule_json["target_sheet"],
                "Target_Header": rule_json["target_header"],
                "Match_Notes": rule_json["match_notes"],
                "Approved": "TRUE",  # default TRUE if missing/blank
                "Rule_ID": "",
                "Processed_At": "",
            }
            if tgt_sheet and tgt_header:
                active_rows.append(row_obj)
            else:
                pending_rows.append(row_obj)
        if pending_rows or active_rows:
            bucket = out.setdefault(company_id, {"pending": [], "active": []})
            if pending_rows:
                bucket["pending"].extend(pending_rows)
            if active_rows:
                bucket["active"].extend(active_rows)
    return out


def rows_to_cells(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    def to_cell(value: Any) -> Dict[str, Any]:
        return {"userEnteredValue": {"stringValue": "" if value is None else str(value)}}
    def to_row(row_vals: List[Any]) -> Dict[str, Any]:
        return {"values": [to_cell(v) for v in row_vals]}
    return [
        to_row([r.get(h, "") for h in PENDING_HEADERS])
        for r in rows
    ]


def active_rows_to_cells(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    # Active header order: Ruleset_Version, Effective_At, Company_ID, Source, Target_Sheet, Target_Header, Match_Notes, Created_At
    from datetime import datetime

    def to_cell(value: Any) -> Dict[str, Any]:
        return {"userEnteredValue": {"stringValue": "" if value is None else str(value)}}

    def to_row(row: Dict[str, Any]) -> Dict[str, Any]:
        now_iso = datetime.utcnow().isoformat(timespec="seconds") + "Z"
        return {
            "values": [
                to_cell(""),                           # Ruleset_Version (blank)
                to_cell(now_iso),                       # Effective_At (seed time)
                to_cell(row.get("Company_ID", "")),
                to_cell(row.get("Source", "")),
                to_cell(row.get("Target_Sheet", "")),
                to_cell(row.get("Target_Header", "")),
                to_cell(row.get("Match_Notes", "")),
                to_cell(now_iso),                       # Created_At
            ]
        }

    return [to_row(r) for r in rows]


def seed_pending(service, spreadsheet_id: str, company_id: str, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    # Ensure setup (tabs/headers/validation/protections) before writing
    ensure_setup_batch(service, spreadsheet_id, company_id, execute=True)

    # Use the new tab naming convention
    titles = (f"{company_id} Pending", f"{company_id} Active")
    res = resolve_sheet_ids(service, spreadsheet_id, titles[0], titles[1])
    pending_id = res.existing_titles_to_ids.get(titles[0], res.pending_id)

    # Build updateCells to write values starting from row 2 (index 1)
    body = {
        "requests": [
            # Clear prior values in Pending from row 2 onward
            {
                "updateCells": {
                    "range": {"sheetId": pending_id, "startRowIndex": 1},
                    "fields": "userEnteredValue",
                }
            },
            # Write new rows
            {
                "updateCells": {
                    "range": {
                        "sheetId": pending_id,
                        "startRowIndex": 1,
                        "startColumnIndex": 0,
                    },
                    "fields": "userEnteredValue",
                    "rows": rows_to_cells(rows),
                }
            },
        ]
    }
    return (
        service.spreadsheets()
        .batchUpdate(spreadsheetId=spreadsheet_id, body=body)
        .execute()
    )


def seed_active(service, spreadsheet_id: str, company_id: str, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    # Ensure setup exists
    ensure_setup_batch(service, spreadsheet_id, company_id, execute=True)

    # Use the new tab naming convention
    titles = (f"{company_id} Pending", f"{company_id} Active")
    res = resolve_sheet_ids(service, spreadsheet_id, titles[0], titles[1])
    active_id = res.existing_titles_to_ids.get(titles[1], res.active_id)

    body = {
        "requests": [
            {"updateCells": {"range": {"sheetId": active_id, "startRowIndex": 1}, "fields": "userEnteredValue"}},
            {
                "updateCells": {
                    "range": {"sheetId": active_id, "startRowIndex": 1, "startColumnIndex": 0},
                    "fields": "userEnteredValue",
                    "rows": active_rows_to_cells(rows),
                }
            },
        ]
    }
    return (
        service.spreadsheets()
        .batchUpdate(spreadsheetId=spreadsheet_id, body=body)
        .execute()
    )


def main() -> None:
    p = argparse.ArgumentParser(description="Seed Pending tabs in Google Sheets from Rules.xlsx")
    p.add_argument("--xlsx", default="Rules.xlsx", help="Path to Rules.xlsx")
    p.add_argument("--config", default=os.path.join("config", "companies.json"))
    p.add_argument("--service-account", help="Path to service_account.json (default: secrets/service_account.json)")
    p.add_argument("--company-id", help="Only seed one company (optional)")
    p.add_argument("--execute", action="store_true", help="Actually write to Sheets (default: dry-run payload print)")
    args = p.parse_args()

    if not os.path.exists(args.xlsx):
        raise SystemExit(f"File not found: {args.xlsx}")
    if not os.path.exists(args.config):
        raise SystemExit(f"Config not found: {args.config}")

    cfg = load_companies_config(args.config)
    companies: List[Dict[str, Any]] = cfg.get("companies", [])
    if args.company_id:
        companies = [c for c in companies if c.get("company_id") == args.company_id]
        if not companies:
            raise SystemExit(f"Company not found in config: {args.company_id}")

    sa_path = args.service_account or os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "secrets", "service_account.json")
    service = _get_service(sa_path)

    to_seed = parse_rules_xlsx(args.xlsx, companies)

    for company in companies:
        company_id = company.get("company_id")
        # Use the rules management workbook, not the company transaction workbook
        # This script is for seeding rule management tabs
        sid = "1mdLWjezU5uj7R3Rp8bTo5AGPuA4yY81bQux4aAV3kec"
        bundle = to_seed.get(company_id, {"pending": [], "active": []})
        pend = bundle.get("pending", [])
        act = bundle.get("active", [])
        print(json.dumps({"company": company_id, "spreadsheet_id": sid, "pending": len(pend), "active": len(act)}, indent=2))
        if args.execute:
            if pend:
                seed_pending(service, sid, company_id, pend)
                print(json.dumps({"company": company_id, "pending_wrote": len(pend)}, indent=2))
            if act:
                seed_active(service, sid, company_id, act)
                print(json.dumps({"company": company_id, "active_wrote": len(act)}, indent=2))


if __name__ == "__main__":
    main()


