from __future__ import annotations

import argparse
import json
import os
import re
import uuid
from dataclasses import dataclass
from hashlib import sha256
from typing import Any, Dict, Iterable, List, Optional, Tuple

import psycopg2
import psycopg2.extras
from openpyxl import load_workbook
from services.common.normalize import normalize_description


# ----------------------------
# Parsing and normalization
# ----------------------------


def _strip(val: Any) -> str:
    return ("" if val is None else str(val)).strip()


def _normalize_header(name: str) -> str:
    n = (name or "").strip().lower()
    for ch in [" ", "-", "\u2013", "\u2014"]:
        n = n.replace(ch, "_")
    while "__" in n:
        n = n.replace("__", "_")
    return n


CANONICAL_COLUMNS = [
    "Source",
    "Target_Sheet",
    "Target_Header",
    "Approved",
]


def _canonicalize_headers(headers: List[str]) -> Dict[str, str]:
    alias_to_canonical: Dict[str, str] = {
        "source": "Source",
        "unique_source": "Source",
        "target_sheet": "Target_Sheet",
        "target_header": "Target_Header",
        "approved": "Approved",
    }
    observed_norm_to_raw: Dict[str, str] = {_normalize_header(h): h for h in headers}
    mapping: Dict[str, str] = {}
    for norm, raw in observed_norm_to_raw.items():
        if norm in alias_to_canonical:
            mapping[alias_to_canonical[norm]] = raw
        if norm == "target_sheet" and "Target_Sheet" not in mapping:
            mapping["Target_Sheet"] = raw
        if norm == "target_header" and "Target_Header" not in mapping:
            mapping["Target_Header"] = raw
        if norm == "unique_source" and "Source" not in mapping:
            mapping["Source"] = raw
    return mapping


def _title_to_company_id(title: str) -> str:
    parts = title.split("–", 1)
    base = parts[1] if len(parts) == 2 else title
    return base.strip()


def _slugify_company(company_id: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "_", company_id.strip().lower())
    slug = re.sub(r"_+", "_", slug).strip("_")
    return slug or "company"


def _bool_from_cell(val: Any) -> bool:
    if isinstance(val, bool):
        return val
    s = _strip(val).lower()
    return s in {"true", "1", "y", "yes"}


def _compute_content_hash(rule_json: Dict[str, Any]) -> str:
    payload = json.dumps(rule_json, sort_keys=True, separators=(",", ":"))
    return sha256(payload.encode("utf-8")).hexdigest()


@dataclass
class PendingRow:
    source: str
    target_sheet: str
    target_header: str
    approved: bool
    content_hash: str


def _row_to_pending(header: List[str], row: Tuple[Any, ...], title_company: str) -> Optional[PendingRow]:
    canon = _canonicalize_headers(header)

    def get(cn: str) -> Any:
        raw = canon.get(cn)
        if raw is None:
            return None
        try:
            idx = header.index(raw)
        except ValueError:
            return None
        return row[idx] if idx < len(row) else None

    source = _strip(get("Source")).lower()
    target_sheet = _strip(get("Target_Sheet"))
    target_header = _strip(get("Target_Header"))
    approved = _bool_from_cell(get("Approved"))

    # Required per spec (minimal schema)
    required = [source, target_sheet, target_header]
    if any(v == "" for v in required):
        return None

    rule_json = {
        "source": source,
        "source_norm": normalize_description(source),
        "target_sheet": target_sheet,
        "target_header": target_header,
    }
    content_hash = _compute_content_hash(rule_json)
    return PendingRow(
        source=source,
        target_sheet=target_sheet,
        target_header=target_header,
        approved=approved,
        content_hash=content_hash,
    )


def parse_workbook(xlsx_path: str, only_company: Optional[str] = None) -> Dict[str, List[PendingRow]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    out: Dict[str, List[PendingRow]] = {}
    for ws in wb.worksheets:
        title_company = _title_to_company_id(ws.title)
        if only_company and title_company != only_company:
            continue
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        canon = _canonicalize_headers(header)
        # Must have the required headers present (minimal)
        if not all(c in canon for c in ["Source", "Target_Sheet", "Target_Header", "Approved"]):
            continue
        pendings: List[PendingRow] = []
        consecutive_blank = 0
        for r in rows[1:]:
            if not r:
                consecutive_blank += 1
                if consecutive_blank > 100:
                    break
                continue
            consecutive_blank = 0
            pr = _row_to_pending(header, r, title_company)
            if pr is None:
                continue
            pendings.append(pr)
        if pendings:
            out.setdefault(title_company, []).extend(pendings)
    return out


# ----------------------------
# Database write path (pending tables)
# ----------------------------


def _ensure_pending_table(conn, table_name: str) -> None:
    ddl = f"""
CREATE TABLE IF NOT EXISTS {table_name} (
  row_id        BIGSERIAL PRIMARY KEY,
  source        TEXT NOT NULL,
  target_sheet  TEXT NOT NULL,
  target_header TEXT NOT NULL,
  approved      BOOLEAN NOT NULL,
  content_hash  CHAR(64) NOT NULL,
  created_at    TIMESTAMPTZ NOT NULL DEFAULT now(),
  UNIQUE (source, content_hash)
);
CREATE INDEX IF NOT EXISTS ix_{_slugify_company(table_name)}_approved ON {table_name}(approved);
"""
    with conn.cursor() as cur:
        cur.execute(ddl)


def _copy_into_temp_and_insert(conn, table_name: str, rows: List[PendingRow]) -> Tuple[int, int]:
    with conn.cursor() as cur:
        cur.execute(
            """
CREATE TEMP TABLE IF NOT EXISTS stg_rules (
  source TEXT,
  target_sheet TEXT,
  target_header TEXT,
  approved BOOLEAN,
  content_hash CHAR(64)
) ON COMMIT DROP;
"""
        )
        copy_sql = "COPY stg_rules (source, target_sheet, target_header, approved, content_hash) FROM STDIN"
        with cur.copy(copy_sql) as cp:
            for r in rows:
                cp.write_row(
                    (
                        r.source,
                        r.target_sheet,
                        r.target_header,
                        bool(r.approved),
                        r.content_hash,
                    )
                )
        cur.execute(
            f"""
INSERT INTO {table_name} (
  source, target_sheet, target_header, approved, content_hash
)
SELECT source, target_sheet, target_header, approved, content_hash
FROM stg_rules
ON CONFLICT (source, content_hash) DO NOTHING;
"""
        )
        affected = cur.rowcount if hasattr(cur, "rowcount") else 0
    return affected, 0


def _execute_values_insert(conn, table_name: str, rows: List[PendingRow]) -> Tuple[int, int]:
    tpl = """
INSERT INTO {table} (
  source, target_sheet, target_header, approved, content_hash
) VALUES %s
ON CONFLICT (source, content_hash) DO NOTHING;
""".format(table=table_name)
    values = [
        (
            r.source,
            r.target_sheet,
            r.target_header,
            bool(r.approved),
            r.content_hash,
        )
        for r in rows
    ]
    with conn.cursor() as cur:
        psycopg2.extras.execute_values(
            cur,
            tpl,
            values,
            template="(%s,%s,%s,%s,%s)",
            page_size=1000,
        )
        affected = cur.rowcount if hasattr(cur, "rowcount") else 0
    return affected, 0


def load_pending_for_company(dsn: str, company_id: str, rows: List[PendingRow], copy_threshold: int = 5000) -> int:
    if not rows:
        return 0
    table_name = f"rules_pending_{_slugify_company(company_id)}"
    inserted = 0
    with psycopg2.connect(dsn) as conn:
        conn.autocommit = False
        _ensure_pending_table(conn, table_name)
        if len(rows) >= copy_threshold:
            ins, _ = _copy_into_temp_and_insert(conn, table_name, rows)
        else:
            ins, _ = _execute_values_insert(conn, table_name, rows)
        conn.commit()
        inserted += int(ins)
    return inserted


# ----------------------------
# CLI entry point
# ----------------------------


def main() -> int:
    p = argparse.ArgumentParser(description="Load Rules.xlsx into per-company rules_pending_<cid> tables")
    p.add_argument("--xlsx", default="Rules.xlsx", help="Path to Rules.xlsx")
    p.add_argument("--company", help="Only load a single company (sheet title)")
    p.add_argument("--dsn-map", help="JSON mapping of company_id -> DSN (overrides KYLO_DB_DSN_MAP)")
    p.add_argument("--dsn-map-file", help="Path to JSON file with company_id -> DSN mapping")
    p.add_argument("--copy-threshold", type=int, default=int(os.environ.get("KYLO_RULES_COPY_THRESHOLD", "5000")))
    args = p.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"File not found: {args.xlsx}")
        return 2

    dsn_map_env = None
    if args.dsn_map_file and os.path.exists(args.dsn_map_file):
        with open(args.dsn_map_file, "r", encoding="utf-8") as f:
            dsn_map_env = f.read()
    dsn_map_env = dsn_map_env or args.dsn_map or os.environ.get("KYLO_DB_DSN_MAP")
    if not dsn_map_env:
        print("KYLO_DB_DSN_MAP not set and --dsn-map not provided")
        return 2
    try:
        dsn_map: Dict[str, str] = json.loads(dsn_map_env)
    except Exception as e:
        print(f"Invalid DSN map: {e}")
        return 2

    parsed = parse_workbook(args.xlsx, args.company)
    total_inserted = 0
    for company_id, rows in parsed.items():
        dsn = dsn_map.get(company_id)
        if not dsn:
            print(f"Skip {company_id}: no DSN in KYLO_DB_DSN_MAP")
            continue
        ins = load_pending_for_company(dsn, company_id, rows, copy_threshold=args.copy_threshold)
        print(f"{company_id}: inserted={ins} (pending)")
        total_inserted += ins

    print(f"Done. total_inserted={total_inserted}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


