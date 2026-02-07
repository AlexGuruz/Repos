#!/usr/bin/env python3
"""
Rule Loader for JGDTRUTH
Loads JGD Truth rules from Excel file
Standalone version that doesn't require database_manager

Author: ScriptHub AI Assistant
Date: 2025-01-XX
"""

import openpyxl
import logging
import sys
import json
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime

class RuleLoader:
    """Loads and manages JGD Truth rules from Excel file"""
    
    def __init__(self, jgd_truth_file: str = None):
        """
        Initialize the rule loader
        
        Args:
            jgd_truth_file: Path to JGD Truth Excel file
        """
        # Use default file if not provided
        if jgd_truth_file is None:
            script_dir = Path(__file__).parent.parent
            # Try common locations
            possible_paths = [
                script_dir / 'JGD Truth Current.xlsx',
                script_dir / 'JGDTruth.xlsx',
                Path('JGD Truth Current.xlsx'),
                Path('JGDTruth.xlsx')
            ]
            
            for path in possible_paths:
                if path.exists():
                    jgd_truth_file = str(path)
                    break
            
            if jgd_truth_file is None:
                jgd_truth_file = 'JGD Truth Current.xlsx'  # Default
        
        self.jgd_truth_file = Path(jgd_truth_file)
        
        # Setup logging
        self.setup_logging()
        
        logging.info("INITIALIZING RULE LOADER")
        logging.info("=" * 50)
    
    def setup_logging(self):
        """Setup logging configuration"""
        log_format = '%(asctime)s - %(levelname)s - %(message)s'
        
        # Create logs directory
        script_dir = Path(__file__).parent.parent
        logs_dir = script_dir / 'logs'
        logs_dir.mkdir(exist_ok=True)
        
        # Configure root logger
        logging.basicConfig(
            level=logging.INFO,
            format=log_format,
            handlers=[
                logging.FileHandler(logs_dir / f'rule_loader_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
                logging.StreamHandler(sys.stdout)
            ]
        )
        
        self.logger = logging.getLogger(__name__)
    
    def load_jgd_truth_rules(self) -> List[Dict]:
        """Load rules from JGD Truth Excel file."""
        try:
            if not self.jgd_truth_file.exists():
                self.logger.error(f"JGD Truth file not found: {self.jgd_truth_file}")
                return []
            
            self.logger.info(f"Loading rules from {self.jgd_truth_file}")
            
            workbook = openpyxl.load_workbook(self.jgd_truth_file)
            all_rules = []
            
            for sheet_name in workbook.sheetnames:
                self.logger.info(f"Processing JGD Truth sheet: {sheet_name}")
                worksheet = workbook[sheet_name]
                
                # Rules start from row 2 (row 1 is header)
                rules_in_sheet = 0
                for row in range(2, worksheet.max_row + 1):
                    source = worksheet.cell(row=row, column=1).value  # Column A
                    target_sheet = worksheet.cell(row=row, column=2).value  # Column B
                    target_header = worksheet.cell(row=row, column=3).value  # Column C
                    
                    if source and str(source).strip():
                        source_clean = str(source).strip()
                        
                        rule_data = {
                            'source': source_clean,
                            'target_sheet': str(target_sheet).strip() if target_sheet else '',
                            'target_header': str(target_header).strip() if target_header else '',
                            'jgd_sheet': sheet_name,
                            'confidence_threshold': 0.7  # Default threshold
                        }
                        
                        all_rules.append(rule_data)
                        rules_in_sheet += 1
                        
                        self.logger.debug(f"  Rule: '{source_clean}' -> '{target_sheet}' / '{target_header}'")
                
                self.logger.info(f"  Loaded {rules_in_sheet} rules from sheet '{sheet_name}'")
            
            workbook.close()
            self.logger.info(f"Successfully loaded {len(all_rules)} total rules from JGD Truth")
            
            return all_rules
            
        except Exception as e:
            self.logger.error(f"Error loading JGD Truth rules: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
            return []
    
    def save_rules_to_json(self, rules: List[Dict], output_file: str = None) -> bool:
        """Save rules to JSON file."""
        try:
            if output_file is None:
                script_dir = Path(__file__).parent.parent
                config_dir = script_dir / 'config'
                config_dir.mkdir(exist_ok=True)
                output_file = config_dir / 'jgd_truth_rules.json'
            else:
                output_file = Path(output_file)
            
            self.logger.info(f"Saving {len(rules)} rules to {output_file}")
            
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(rules, f, indent=2, ensure_ascii=False)
            
            self.logger.info(f"Successfully saved rules to {output_file}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error saving rules to JSON: {e}")
            return False
    
    def load_rules_from_json(self, input_file: str = None) -> List[Dict]:
        """Load rules from JSON file."""
        try:
            if input_file is None:
                script_dir = Path(__file__).parent.parent
                config_dir = script_dir / 'config'
                input_file = config_dir / 'jgd_truth_rules.json'
            else:
                input_file = Path(input_file)
            
            if not input_file.exists():
                self.logger.warning(f"JSON file not found: {input_file}")
                return []
            
            self.logger.info(f"Loading rules from {input_file}")
            
            with open(input_file, 'r', encoding='utf-8') as f:
                rules = json.load(f)
            
            self.logger.info(f"Successfully loaded {len(rules)} rules from JSON")
            return rules
            
        except Exception as e:
            self.logger.error(f"Error loading rules from JSON: {e}")
            return []
    
    def validate_rules(self, rules: List[Dict]) -> Dict:
        """Validate loaded rules for completeness."""
        validation_results = {
            'total_rules': len(rules),
            'valid_rules': 0,
            'invalid_rules': 0,
            'missing_target_sheet': 0,
            'missing_target_header': 0,
            'empty_source': 0
        }
        
        for rule in rules:
            is_valid = True
            
            if not rule.get('source'):
                validation_results['empty_source'] += 1
                is_valid = False
            
            if not rule.get('target_sheet'):
                validation_results['missing_target_sheet'] += 1
                is_valid = False
            
            if not rule.get('target_header'):
                validation_results['missing_target_header'] += 1
                is_valid = False
            
            if is_valid:
                validation_results['valid_rules'] += 1
            else:
                validation_results['invalid_rules'] += 1
        
        return validation_results
    
    def get_rule_statistics(self, rules: List[Dict] = None) -> Dict:
        """Get statistics about loaded rules."""
        try:
            if rules is None:
                rules = self.load_jgd_truth_rules()
            
            stats = {
                'total_rules': len(rules),
                'unique_sources': len(set(rule['source'] for rule in rules)),
                'unique_target_sheets': len(set(rule['target_sheet'] for rule in rules)),
                'unique_target_headers': len(set(rule['target_header'] for rule in rules))
            }
            
            # Count rules by target sheet
            sheet_counts = {}
            for rule in rules:
                sheet = rule['target_sheet']
                sheet_counts[sheet] = sheet_counts.get(sheet, 0) + 1
            
            stats['rules_by_sheet'] = sheet_counts
            
            return stats
            
        except Exception as e:
            self.logger.error(f"Error getting rule statistics: {e}")
            return {}

def main():
    """Test rule loader."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Load and validate JGD Truth rules')
    parser.add_argument('--file', '-f', 
                       help='Path to JGD Truth Excel file')
    parser.add_argument('--save-json', '-s', action='store_true',
                       help='Save rules to JSON file')
    parser.add_argument('--verbose', '-v', action='store_true',
                       help='Enable verbose logging')
    
    args = parser.parse_args()
    
    print("RULE LOADER TEST")
    print("=" * 50)
    
    rule_loader = RuleLoader(jgd_truth_file=args.file)
    
    # Load rules from JGD Truth file
    rules = rule_loader.load_jgd_truth_rules()
    
    if rules:
        print(f"✅ Successfully loaded {len(rules)} rules from JGD Truth file")
        
        # Validate rules
        validation = rule_loader.validate_rules(rules)
        print(f"\n📊 Rule Validation:")
        print(f"  Total rules: {validation['total_rules']}")
        print(f"  Valid rules: {validation['valid_rules']}")
        print(f"  Invalid rules: {validation['invalid_rules']}")
        
        if validation['invalid_rules'] > 0:
            print(f"  Missing target sheet: {validation['missing_target_sheet']}")
            print(f"  Missing target header: {validation['missing_target_header']}")
            print(f"  Empty source: {validation['empty_source']}")
        
        # Show first few rules
        print(f"\nFirst 5 rules:")
        for i, rule in enumerate(rules[:5]):
            print(f"  {i+1}. '{rule['source']}' → '{rule['target_sheet']}' / '{rule['target_header']}'")
        
        # Save to JSON if requested
        if args.save_json:
            if rule_loader.save_rules_to_json(rules):
                print(f"\n✅ Successfully saved rules to JSON")
            else:
                print(f"\n❌ Failed to save rules to JSON")
        
        # Get statistics
        stats = rule_loader.get_rule_statistics(rules)
        print(f"\n📊 Statistics:")
        print(f"  Total rules: {stats['total_rules']}")
        print(f"  Unique sources: {stats['unique_sources']}")
        print(f"  Unique target sheets: {stats['unique_target_sheets']}")
        print(f"  Unique target headers: {stats['unique_target_headers']}")
        
        # Show rules by sheet
        print(f"\nRules by target sheet:")
        for sheet, count in stats['rules_by_sheet'].items():
            print(f"  {sheet}: {count} rules")
        
    else:
        print("❌ Failed to load rules from JGD Truth file")

if __name__ == "__main__":
    main()

