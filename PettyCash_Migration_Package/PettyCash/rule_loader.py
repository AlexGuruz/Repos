#!/usr/bin/env python3
"""
Rule Loader for Petty Cash Sorter
Loads JGD Truth rules from Excel file into database and AI matcher
"""

import openpyxl
import logging
import sqlite3
from pathlib import Path
from typing import List, Dict, Optional
from database_manager import DatabaseManager

class RuleLoader:
    def __init__(self, jgd_truth_file: str = "JGD Truth Current.xlsx"):
        self.jgd_truth_file = Path(jgd_truth_file)
        self.db_manager = DatabaseManager()
        
        # Configure logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('logs/rule_loader.log'),
                logging.StreamHandler()
            ]
        )
        
        logging.info("INITIALIZING RULE LOADER")
        logging.info("=" * 50)
    
    def load_jgd_truth_rules(self) -> List[Dict]:
        """Load rules from JGD Truth Excel file."""
        try:
            if not self.jgd_truth_file.exists():
                logging.error(f"JGD Truth file not found: {self.jgd_truth_file}")
                return []
            
            logging.info(f"Loading rules from {self.jgd_truth_file}")
            
            workbook = openpyxl.load_workbook(self.jgd_truth_file)
            all_rules = []
            
            for sheet_name in workbook.sheetnames:
                logging.info(f"Processing JGD Truth sheet: {sheet_name}")
                worksheet = workbook[sheet_name]
                
                # Rules start from row 2 (row 1 is header)
                rules_in_sheet = 0
                for row in range(2, worksheet.max_row + 1):
                    source = worksheet.cell(row=row, column=1).value  # Column A
                    target_sheet = worksheet.cell(row=row, column=2).value  # Column B
                    target_header = worksheet.cell(row=row, column=3).value  # Column C
                    
                    if source and str(source).strip():
                        source_clean = str(source).strip()
                        
                        rule_data = {
                            'source': source_clean,
                            'target_sheet': str(target_sheet).strip() if target_sheet else '',
                            'target_header': str(target_header).strip() if target_header else '',
                            'jgd_sheet': sheet_name,
                            'confidence_threshold': 0.7  # Default threshold
                        }
                        
                        all_rules.append(rule_data)
                        rules_in_sheet += 1
                        
                        logging.debug(f"  Rule: '{source_clean}' -> '{target_sheet}' / '{target_header}'")
                
                logging.info(f"  Loaded {rules_in_sheet} rules from sheet '{sheet_name}'")
            
            workbook.close()
            logging.info(f"Successfully loaded {len(all_rules)} total rules from JGD Truth")
            
            return all_rules
            
        except Exception as e:
            logging.error(f"Error loading JGD Truth rules: {e}")
            return []
    
    def save_rules_to_database(self, rules: List[Dict]) -> bool:
        """Save rules to SQLite database."""
        try:
            logging.info("Saving rules to database...")
            
            success_count = 0
            for rule in rules:
                if self.db_manager.add_rule(rule):
                    success_count += 1
                else:
                    logging.warning(f"Failed to save rule: {rule['source']}")
            
            logging.info(f"Successfully saved {success_count}/{len(rules)} rules to database")
            return success_count == len(rules)
            
        except Exception as e:
            logging.error(f"Error saving rules to database: {e}")
            return False
    
    def get_rules_from_database(self) -> List[Dict]:
        """Get all rules from database."""
        return self.db_manager.get_all_rules()
    
    def clear_database_rules(self) -> bool:
        """Clear all rules from database."""
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cursor = conn.cursor()
            
            cursor.execute('DELETE FROM rules')
            conn.commit()
            conn.close()
            
            logging.info("Cleared all rules from database")
            return True
            
        except Exception as e:
            logging.error(f"Error clearing database rules: {e}")
            return False
    
    def reload_rules(self) -> bool:
        """Reload all rules from JGD Truth file to database."""
        try:
            logging.info("Reloading rules from JGD Truth file...")
            
            # Clear existing rules
            if not self.clear_database_rules():
                return False
            
            # Load rules from Excel
            rules = self.load_jgd_truth_rules()
            if not rules:
                logging.error("No rules loaded from JGD Truth file")
                return False
            
            # Save to database
            if not self.save_rules_to_database(rules):
                return False
            
            logging.info("Successfully reloaded all rules")
            return True
            
        except Exception as e:
            logging.error(f"Error reloading rules: {e}")
            return False
    
    def validate_rules(self, rules: List[Dict]) -> Dict:
        """Validate loaded rules for completeness."""
        validation_results = {
            'total_rules': len(rules),
            'valid_rules': 0,
            'invalid_rules': 0,
            'missing_target_sheet': 0,
            'missing_target_header': 0,
            'empty_source': 0
        }
        
        for rule in rules:
            is_valid = True
            
            if not rule.get('source'):
                validation_results['empty_source'] += 1
                is_valid = False
            
            if not rule.get('target_sheet'):
                validation_results['missing_target_sheet'] += 1
                is_valid = False
            
            if not rule.get('target_header'):
                validation_results['missing_target_header'] += 1
                is_valid = False
            
            if is_valid:
                validation_results['valid_rules'] += 1
            else:
                validation_results['invalid_rules'] += 1
        
        return validation_results
    
    def get_rule_statistics(self) -> Dict:
        """Get statistics about loaded rules."""
        try:
            rules = self.get_rules_from_database()
            
            stats = {
                'total_rules': len(rules),
                'unique_sources': len(set(rule['source'] for rule in rules)),
                'unique_target_sheets': len(set(rule['target_sheet'] for rule in rules)),
                'unique_target_headers': len(set(rule['target_header'] for rule in rules))
            }
            
            # Count rules by target sheet
            sheet_counts = {}
            for rule in rules:
                sheet = rule['target_sheet']
                sheet_counts[sheet] = sheet_counts.get(sheet, 0) + 1
            
            stats['rules_by_sheet'] = sheet_counts
            
            return stats
            
        except Exception as e:
            logging.error(f"Error getting rule statistics: {e}")
            return {}

def main():
    """Test rule loader."""
    print("RULE LOADER TEST")
    print("=" * 50)
    
    rule_loader = RuleLoader()
    
    # Load rules from JGD Truth file
    rules = rule_loader.load_jgd_truth_rules()
    
    if rules:
        print(f"✅ Successfully loaded {len(rules)} rules from JGD Truth file")
        
        # Validate rules
        validation = rule_loader.validate_rules(rules)
        print(f"\n📊 Rule Validation:")
        print(f"  Total rules: {validation['total_rules']}")
        print(f"  Valid rules: {validation['valid_rules']}")
        print(f"  Invalid rules: {validation['invalid_rules']}")
        
        if validation['invalid_rules'] > 0:
            print(f"  Missing target sheet: {validation['missing_target_sheet']}")
            print(f"  Missing target header: {validation['missing_target_header']}")
            print(f"  Empty source: {validation['empty_source']}")
        
        # Show first few rules
        print(f"\nFirst 5 rules:")
        for i, rule in enumerate(rules[:5]):
            print(f"  {i+1}. '{rule['source']}' → '{rule['target_sheet']}' / '{rule['target_header']}'")
        
        # Save to database
        if rule_loader.save_rules_to_database(rules):
            print(f"\n✅ Successfully saved rules to database")
            
            # Get statistics
            stats = rule_loader.get_rule_statistics()
            print(f"\n📊 Database Statistics:")
            print(f"  Total rules: {stats['total_rules']}")
            print(f"  Unique sources: {stats['unique_sources']}")
            print(f"  Unique target sheets: {stats['unique_target_sheets']}")
            print(f"  Unique target headers: {stats['unique_target_headers']}")
            
            # Show rules by sheet
            print(f"\nRules by target sheet:")
            for sheet, count in stats['rules_by_sheet'].items():
                print(f"  {sheet}: {count} rules")
        
        else:
            print(f"\n❌ Failed to save rules to database")
        
    else:
        print("❌ Failed to load rules from JGD Truth file")

if __name__ == "__main__":
    main() 