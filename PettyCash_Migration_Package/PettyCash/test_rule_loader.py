#!/usr/bin/env python3
"""
Simple test for rule loader
"""

import openpyxl
from pathlib import Path

def test_jgd_truth_access():
    """Test if we can access the JGD Truth file."""
    print("TESTING JGD TRUTH ACCESS")
    print("=" * 50)
    
    jgd_file = Path("JGD Truth Current.xlsx")
    
    if not jgd_file.exists():
        print(f"❌ JGD Truth file not found: {jgd_file}")
        return False
    
    print(f"✅ Found JGD Truth file: {jgd_file}")
    
    try:
        workbook = openpyxl.load_workbook(jgd_file)
        print(f"✅ Successfully opened workbook")
        print(f"📋 Sheets: {workbook.sheetnames}")
        
        # Test first sheet
        if workbook.sheetnames:
            first_sheet = workbook.sheetnames[0]
            worksheet = workbook[first_sheet]
            print(f"📊 First sheet '{first_sheet}' has {worksheet.max_row} rows and {worksheet.max_column} columns")
            
            # Test reading first few rows
            print(f"\nFirst 3 rows from sheet '{first_sheet}':")
            for row in range(1, min(4, worksheet.max_row + 1)):
                row_data = []
                for col in range(1, min(4, worksheet.max_column + 1)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value else '')
                print(f"  Row {row}: {row_data}")
        
        workbook.close()
        return True
        
    except Exception as e:
        print(f"❌ Error accessing JGD Truth file: {e}")
        return False

if __name__ == "__main__":
    test_jgd_truth_access() 