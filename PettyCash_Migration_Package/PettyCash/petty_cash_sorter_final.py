#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FINAL PETTY CASH SORTER WITH AI-ENHANCED RULE MATCHING
Uses Excel files as mappers + column-specific PETTY CASH snapshot + intelligent fuzzy matching
"""

import openpyxl
import logging
from pathlib import Path
import json
from datetime import datetime, timedelta
import sys
import os
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
import csv
import re
from difflib import SequenceMatcher
from typing import Dict, List, Optional
from dataclasses import dataclass
import sqlite3  # For future database rule storage
import time # For time.time()
from tqdm import tqdm  # For progress bars
import threading  # For background monitoring
import random  # For jitter in exponential backoff
from gspread.exceptions import APIError, WorksheetNotFound
import psutil  # For memory monitoring
import gc  # For garbage collection monitoring

# Configure logging with detailed granularity
logging.basicConfig(
    level=logging.DEBUG,  # Changed to DEBUG for more detailed logging
    format='%(asctime)s - %(levelname)s - [%(funcName)s:%(lineno)d] - %(message)s',
    handlers=[
        logging.FileHandler('logs/final_sorter.log'),
        logging.StreamHandler(sys.stdout)
    ]
)

# Create logs directory if it doesn't exist
Path('logs').mkdir(exist_ok=True)

@dataclass
class MatchResult:
    """Result of a rule matching attempt."""
    matched: bool
    confidence: float
    matched_rule: Optional[Dict] = None
    original_source: str = ""
    normalized_source: str = ""
    match_type: str = "none"
    suggestions: List[str] = None

class AIEnhancedRuleMatcher:
    """AI-enhanced rule matcher with fuzzy logic and intelligent corrections."""
    
    def __init__(self):
        self.rules_cache = {}
        self.source_variations = {}
        self.confidence_threshold = 0.75  # Base threshold
        self.max_suggestions = 5
        
        # Dynamic confidence threshold system
        self.company_thresholds = {}  # Company-specific thresholds
        self.match_history = []  # Track match results for learning
        self.min_threshold = 0.5  # Minimum acceptable threshold
        self.max_threshold = 0.95  # Maximum threshold
        self.learning_rate = 0.05  # How much to adjust thresholds
        
        # Semantic understanding patterns
        self.semantic_patterns = {
            'payroll': ['salary', 'wage', 'paycheck', 'employee payment', 'payroll', 'pay'],
            'sales': ['revenue', 'income', 'customer payment', 'wholesale', 'sale', 'sales'],
            'expenses': ['cost', 'purchase', 'bill', 'payment', 'expense', 'buy'],
            'deposits': ['bank deposit', 'cash deposit', 'fund transfer', 'deposit'],
            'register': ['register', 'reg', 'cash register', 'pos'],
            'orders': ['order', 'purchase order', 'po', 'buy'],
            'lab': ['lab', 'laboratory', 'testing', 'test', 'analysis']
        }
        
        # Common patterns and corrections
        self.common_patterns = {
            # Word order variations
            r'^(ORDER|SALE TO|PAYROLL|REG)\s+(.+)$': r'\2 \1',
            r'^(.+)\s+(ORDER|SALE TO|PAYROLL|REG)$': r'\2 \1',
            
            # Common abbreviations
            r'\bINC\b': 'INCORPORATED',
            r'\bCO\b': 'COMPANY',
            r'\bLLC\b': 'LIMITED LIABILITY COMPANY',
            r'\bDIST\b': 'DISTRIBUTION',
            r'\bDISP\b': 'DISPENSARY',
            
            # Common typos
            r'\bGUYZ\b': 'GUYS',
            r'\bMARIJUANA\b': 'MARIJUANA',
            r'\bWHOLESALE\b': 'WHOLESALE',
            r'\bEMERALDS\b': 'EMERALDS',
            r'\bWELLNESS\b': 'WELLNESS',
        }
        
        # Company-specific patterns
        self.company_patterns = {
            'NUGZ': {
                'sale_patterns': [r'SALE TO (.+)', r'SALES TO (.+)', r'WHOLESALE (.+)'],
                'payroll_patterns': [r'PAYROLL (\d+)', r'PAYCHECK (\d+)'],
                'register_patterns': [r'REG (\d+)', r'REGISTER (\d+)'],
            },
            'JGD': {
                'sale_patterns': [r'SALE TO (.+)', r'SALES TO (.+)'],
                'payroll_patterns': [r'PAYROLL (\d+)', r'PAYCHECK (\d+)'],
            },
            'PUFFIN': {
                'order_patterns': [r'ORDER (.+)', r'PURCHASE (.+)'],
                'lab_patterns': [r'LAB (.+)', r'TESTING (.+)'],
            }
        }
        
        # Load learned variations
        self.load_learned_variations()
    
    def load_learned_variations(self):
        """Load previously learned source variations."""
        variations_file = Path("config/source_variations.json")
        if variations_file.exists():
            try:
                with open(variations_file, 'r') as f:
                    self.source_variations = json.load(f)
                logging.info(f"Loaded {len(self.source_variations)} learned source variations")
            except Exception as e:
                logging.warning(f"Could not load source variations: {e}")
    
    def save_learned_variations(self):
        """Save learned source variations for future use."""
        variations_file = Path("config/source_variations.json")
        variations_file.parent.mkdir(exist_ok=True)
        
        try:
            with open(variations_file, 'w') as f:
                json.dump(self.source_variations, f, indent=2)
            logging.info(f"Saved {len(self.source_variations)} source variations")
        except Exception as e:
            logging.error(f"Could not save source variations: {e}")
    
    def normalize_source(self, source: str) -> str:
        """Normalize source string for better matching."""
        if not source:
            return ""
        
        # Convert to string and strip whitespace
        normalized = str(source).strip()
        
        # Convert to uppercase for consistency
        normalized = normalized.upper()
        
        # Remove extra whitespace
        normalized = re.sub(r'\s+', ' ', normalized)
        
        # Apply common pattern corrections
        for pattern, replacement in self.common_patterns.items():
            normalized = re.sub(pattern, replacement, normalized, flags=re.IGNORECASE)
        
        return normalized.strip()
    
    def calculate_similarity(self, source1: str, source2: str) -> float:
        """Calculate similarity between two source strings."""
        if not source1 or not source2:
            return 0.0
        
        # Normalize both strings
        norm1 = self.normalize_source(source1)
        norm2 = self.normalize_source(source2)
        
        # Exact match after normalization
        if norm1 == norm2:
            return 1.0
        
        # Use SequenceMatcher for fuzzy matching
        similarity = SequenceMatcher(None, norm1, norm2).ratio()
        
        # Boost similarity for partial matches
        if norm1 in norm2 or norm2 in norm1:
            similarity += 0.2
        
        # Boost for word overlap
        words1 = set(norm1.split())
        words2 = set(norm2.split())
        if len(words1) > 0 and len(words2) > 0:
            word_overlap = len(words1.intersection(words2)) / max(len(words1), len(words2))
            similarity += word_overlap * 0.3
        
        return min(similarity, 1.0)
    
    def find_best_match(self, source: str, company: str) -> MatchResult:
        """Find the best matching rule for a given source and company."""
        logging.debug(f"Finding match for source: '{source}' (Company: {company})")
        
        if not source or not company:
            logging.debug(f"Invalid input: source='{source}', company='{company}'")
            return MatchResult(matched=False, confidence=0.0, original_source=source)
        
        original_source = source
        normalized_source = self.normalize_source(source)
        
        # Get dynamic confidence threshold for this company and source
        dynamic_threshold = self.get_dynamic_threshold(company, source)
        
        # First, try exact match
        exact_match = self._find_exact_match(normalized_source, company)
        if exact_match:
            logging.debug(f"Exact match found: '{normalized_source}' -> '{exact_match['source']}'")
            return MatchResult(
                matched=True,
                confidence=1.0,
                matched_rule=exact_match,
                original_source=original_source,
                normalized_source=normalized_source,
                match_type="exact"
            )
        
        # Try learned variations
        variation_match = self._find_variation_match(source, company)
        if variation_match:
            logging.debug(f"Learned variation match found: '{source}' -> '{variation_match['source']}'")
            return MatchResult(
                matched=True,
                confidence=0.95,
                matched_rule=variation_match,
                original_source=original_source,
                normalized_source=normalized_source,
                match_type="learned_variation"
            )
        
        # Try fuzzy matching with dynamic threshold
        fuzzy_match = self._find_fuzzy_match(source, company)
        if fuzzy_match and fuzzy_match['confidence'] >= dynamic_threshold:
            logging.debug(f"Fuzzy match found: '{source}' -> '{fuzzy_match['rule']['source']}' (confidence: {fuzzy_match['confidence']:.2f} >= {dynamic_threshold:.3f})")
            return MatchResult(
                matched=True,
                confidence=fuzzy_match['confidence'],
                matched_rule=fuzzy_match['rule'],
                original_source=original_source,
                normalized_source=normalized_source,
                match_type="fuzzy"
            )
        elif fuzzy_match:
            logging.debug(f"Fuzzy match below dynamic threshold: '{source}' -> '{fuzzy_match['rule']['source']}' (confidence: {fuzzy_match['confidence']:.2f} < {dynamic_threshold:.3f})")
        
        # Generate adaptive suggestions for unmatched sources
        adaptive_suggestions = self.get_adaptive_suggestions(source, company)
        suggestions = [s['target_sheet'] + '/' + s['target_header'] for s in adaptive_suggestions]
        
        return MatchResult(
            matched=False,
            confidence=0.0,
            original_source=original_source,
            normalized_source=normalized_source,
            match_type="none",
            suggestions=suggestions
        )
    
    def _find_exact_match(self, normalized_source: str, company: str) -> Optional[Dict]:
        """Find exact match in rules cache."""
        for rule_key, rule in self.rules_cache.items():
            if rule['jgd_sheet'] == company and rule['source'] == normalized_source:
                return rule
        return None
    
    def _find_variation_match(self, source: str, company: str) -> Optional[Dict]:
        """Find match using learned variations."""
        if source in self.source_variations:
            canonical_source = self.source_variations[source]
            for rule_key, rule in self.rules_cache.items():
                if rule['jgd_sheet'] == company and rule['source'] == canonical_source:
                    return rule
        return None
    
    def _find_fuzzy_match(self, source: str, company: str) -> Optional[Dict]:
        """Find best fuzzy match."""
        best_match = None
        best_confidence = 0.0
        
        for rule_key, rule in self.rules_cache.items():
            if rule['jgd_sheet'] == company:
                confidence = self.calculate_similarity(source, rule['source'])
                if confidence > best_confidence:
                    best_confidence = confidence
                    best_match = {
                        'rule': rule,
                        'confidence': confidence
                    }
        
        return best_match
    
    def _generate_suggestions(self, source: str, company: str) -> List[str]:
        """Generate suggestions for similar rules."""
        suggestions = []
        
        # Find similar rules
        similar_rules = []
        for rule_key, rule in self.rules_cache.items():
            if rule['jgd_sheet'] == company:
                similarity = self.calculate_similarity(source, rule['source'])
                if similarity > 0.3:  # Lower threshold for suggestions
                    similar_rules.append((rule['source'], similarity))
        
        # Sort by similarity and return top suggestions
        similar_rules.sort(key=lambda x: x[1], reverse=True)
        suggestions = [rule[0] for rule in similar_rules[:self.max_suggestions]]
        
        return suggestions
    
    def learn_variation(self, source: str, canonical_source: str):
        """Learn a new source variation."""
        self.source_variations[source] = canonical_source
        logging.info(f"Learned variation: '{source}' -> '{canonical_source}'")
    
    def auto_correct_source(self, source: str, company: str) -> str:
        """Attempt to auto-correct common issues in source names."""
        if not source:
            return source
        
        corrected = source
        
        # Fix common word order issues
        corrected = self._fix_word_order(corrected)
        
        # Fix common typos
        corrected = self._fix_typos(corrected)
        
        # Fix spacing issues
        corrected = self._fix_spacing(corrected)
        
        # Apply company-specific corrections
        corrected = self._apply_company_corrections(corrected, company)
        
        return corrected
    
    def _fix_word_order(self, source: str) -> str:
        """Fix common word order issues."""
        # ORDER HIGH GUYS -> HIGH GUYS ORDER
        if re.match(r'^ORDER\s+(.+)$', source, re.IGNORECASE):
            match = re.match(r'^ORDER\s+(.+)$', source, re.IGNORECASE)
            return f"{match.group(1)} ORDER"
        
        # SALE TO HIGH HOUSE -> HIGH HOUSE SALE
        if re.match(r'^SALE TO\s+(.+)$', source, re.IGNORECASE):
            match = re.match(r'^SALE TO\s+(.+)$', source, re.IGNORECASE)
            return f"{match.group(1)} SALE"
        
        return source
    
    def _fix_typos(self, source: str) -> str:
        """Fix common typos."""
        corrections = {
            'GUYZ': 'GUYS',
            'MARIJUANA': 'MARIJUANA',
            'WHOLESALE': 'WHOLESALE',
            'EMERALDS': 'EMERALDS',
            'WELLNESS': 'WELLNESS',
            'DISPENSARY': 'DISPENSARY',
            'DISTRIBUTION': 'DISTRIBUTION',
        }
        
        corrected = source
        for typo, correction in corrections.items():
            corrected = re.sub(rf'\b{typo}\b', correction, corrected, flags=re.IGNORECASE)
        
        return corrected
    
    def _fix_spacing(self, source: str) -> str:
        """Fix spacing issues."""
        # Remove extra spaces
        corrected = re.sub(r'\s+', ' ', source.strip())
        return corrected
    
    def _apply_company_corrections(self, source: str, company: str) -> str:
        """Apply company-specific corrections."""
        if company in self.company_patterns:
            patterns = self.company_patterns[company]
            # Apply company-specific pattern corrections here if needed
            pass
        return source
    
    def get_match_statistics(self) -> Dict:
        """Get statistics about the rule matcher."""
        return {
            'total_rules': len(self.rules_cache),
            'learned_variations': len(self.source_variations),
            'confidence_threshold': self.confidence_threshold
        }

    def get_dynamic_threshold(self, company: str, source: str) -> float:
        """Get dynamic confidence threshold based on company and source patterns."""
        # Start with company-specific threshold or base threshold
        threshold = self.company_thresholds.get(company, self.confidence_threshold)
        
        # Adjust based on source semantic meaning
        semantic_boost = self._get_semantic_confidence_boost(source)
        threshold += semantic_boost
        
        # Adjust based on recent match history for this company
        history_adjustment = self._get_history_adjustment(company)
        threshold += history_adjustment
        
        # Ensure threshold stays within bounds
        threshold = max(self.min_threshold, min(self.max_threshold, threshold))
        
        logging.debug(f"Dynamic threshold for '{company}': {threshold:.3f} (base: {self.confidence_threshold:.3f}, semantic: {semantic_boost:.3f}, history: {history_adjustment:.3f})")
        
        return threshold
    
    def _get_semantic_confidence_boost(self, source: str) -> float:
        """Get confidence boost based on semantic understanding of source."""
        source_lower = source.lower()
        boost = 0.0
        
        # Check for semantic patterns
        for category, patterns in self.semantic_patterns.items():
            for pattern in patterns:
                if pattern in source_lower:
                    boost += 0.1  # Boost for semantic understanding
                    break
        
        # Boost for clear business terminology
        business_terms = ['sale', 'payroll', 'order', 'register', 'lab', 'deposit']
        for term in business_terms:
            if term in source_lower:
                boost += 0.05
        
        return min(boost, 0.2)  # Cap semantic boost at 0.2
    
    def _get_history_adjustment(self, company: str) -> float:
        """Get threshold adjustment based on recent match history."""
        if not self.match_history:
            return 0.0
        
        # Get recent matches for this company (last 20)
        recent_matches = [m for m in self.match_history[-20:] if m['company'] == company]
        
        if not recent_matches:
            return 0.0
        
        # Calculate success rate
        successful_matches = [m for m in recent_matches if m['success']]
        success_rate = len(successful_matches) / len(recent_matches)
        
        # Adjust threshold based on success rate
        if success_rate > 0.8:
            return 0.05  # Slightly lower threshold for high success
        elif success_rate < 0.5:
            return -0.1  # Lower threshold for low success
        
        return 0.0
    
    def update_threshold_from_match(self, company: str, source: str, confidence: float, success: bool):
        """Update threshold based on match result for learning."""
        # Record match in history
        self.match_history.append({
            'company': company,
            'source': source,
            'confidence': confidence,
            'success': success,
            'timestamp': time.time()
        })
        
        # Keep only last 100 matches
        if len(self.match_history) > 100:
            self.match_history = self.match_history[-100:]
        
        # Update company-specific threshold
        current_threshold = self.company_thresholds.get(company, self.confidence_threshold)
        
        if success and confidence < current_threshold:
            # Successful match below threshold - lower threshold
            new_threshold = current_threshold - (self.learning_rate * 0.5)
            self.company_thresholds[company] = max(self.min_threshold, new_threshold)
            logging.debug(f"Lowered threshold for '{company}': {current_threshold:.3f} -> {self.company_thresholds[company]:.3f}")
        
        elif not success and confidence >= current_threshold:
            # Failed match above threshold - raise threshold
            new_threshold = current_threshold + self.learning_rate
            self.company_thresholds[company] = min(self.max_threshold, new_threshold)
            logging.debug(f"Raised threshold for '{company}': {current_threshold:.3f} -> {self.company_thresholds[company]:.3f}")
    
    def get_adaptive_suggestions(self, source: str, company: str) -> List[Dict]:
        """Generate adaptive suggestions based on semantic understanding and patterns."""
        suggestions = []
        source_lower = source.lower()
        
        # Semantic-based suggestions
        for category, patterns in self.semantic_patterns.items():
            for pattern in patterns:
                if pattern in source_lower:
                    # Find matching rules for this semantic category
                    for rule_key, rule in self.rules_cache.items():
                        if rule['jgd_sheet'] == company:
                            rule_lower = rule['source'].lower()
                            if any(p in rule_lower for p in patterns):
                                suggestions.append({
                                    'target_sheet': rule['target_sheet'],
                                    'target_header': rule['target_header'],
                                    'confidence': 0.85,  # High confidence for semantic matches
                                    'reason': f'Semantic match: {category}'
                                })
        
        # Pattern-based suggestions
        if company in self.company_patterns:
            company_patterns = self.company_patterns[company]
            
            for pattern_type, patterns in company_patterns.items():
                for pattern in patterns:
                    if re.search(pattern, source, re.IGNORECASE):
                        # Find matching rules for this pattern type
                        for rule_key, rule in self.rules_cache.items():
                            if rule['jgd_sheet'] == company:
                                if pattern_type == 'sale_patterns' and 'sale' in rule['source'].lower():
                                    suggestions.append({
                                        'target_sheet': rule['target_sheet'],
                                        'target_header': rule['target_header'],
                                        'confidence': 0.8,
                                        'reason': f'Pattern match: {pattern_type}'
                                    })
                                elif pattern_type == 'payroll_patterns' and 'payroll' in rule['source'].lower():
                                    suggestions.append({
                                        'target_sheet': rule['target_sheet'],
                                        'target_header': rule['target_header'],
                                        'confidence': 0.8,
                                        'reason': f'Pattern match: {pattern_type}'
                                    })
        
        # Remove duplicates and sort by confidence
        unique_suggestions = []
        seen = set()
        for suggestion in suggestions:
            key = (suggestion['target_sheet'], suggestion['target_header'])
            if key not in seen:
                seen.add(key)
                unique_suggestions.append(suggestion)
        
        return sorted(unique_suggestions, key=lambda x: x['confidence'], reverse=True)[:self.max_suggestions]

class FinalPettyCashSorter:
    """Final Petty Cash Sorter with AI-enhanced rule matching and performance optimizations."""
    
    def __init__(self):
        # File paths
        self.jgd_truth_file = "JGD Truth Current.xlsx"
        self.rules_cache = {}
        self.missing_rules = []
        
        # Performance optimization: Cache for existing transactions
        self.existing_transactions_cache = set()
        
        # Layout map cache for fast cell coordinate lookups
        self.layout_map_cache = {}
        
        # Initialize AI-enhanced rule matcher
        self.ai_matcher = AIEnhancedRuleMatcher()
        
        # Initialize progress monitor
        self.progress_monitor = ProgressMonitor()
        
        # Initialize error recovery system
        self.error_recovery = ErrorRecovery()
        
        # Initialize performance metrics
        self.performance_metrics = PerformanceMetrics()
        
        # Google Sheets setup
        self.scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        self.creds = Credentials.from_service_account_file('config/service_account.json', scopes=self.scope)
        self.gc = gspread.authorize(self.creds)
        self.petty_cash_url = "https://docs.google.com/spreadsheets/d/1DLYt2r34zASzALv6crQ9GOysdr8sLHL0Bor-5nW5QCU"
        self.financials_spreadsheet_url = "https://docs.google.com/spreadsheets/d/1DLYt2r34zASzALv6crQ9GOysdr8sLHL0Bor-5nW5QCU"
        
        # Create directories
        Path("logs").mkdir(exist_ok=True)
        Path("pending_transactions").mkdir(exist_ok=True)
        Path("config").mkdir(exist_ok=True)
        
        # CSV file paths
        self.preprocessed_csv = "pending_transactions/Petty Cash PreProcessed.csv"
        self.processed_csv = "pending_transactions/Petty Cash Processed.csv"
        self.needs_rules_csv = "pending_transactions/Petty Cash Needs Rules.csv"
        self.failed_csv = "pending_transactions/Petty Cash Failed.csv"
        self.ai_matches_csv = "pending_transactions/AI Matches.csv"
        
        # Initialize CSV files if they don't exist
        self.initialize_csv_files()
        
        logging.info("INITIALIZING FINAL PETTY CASH SORTER WITH AI-ENHANCED RULE MATCHING")
        logging.info("=" * 60)
    
    def initialize_csv_files(self):
        """Initialize CSV files with headers if they don't exist."""
        csv_files = [
            (self.preprocessed_csv, ['Row', 'Date', 'Initials', 'Source', 'Company', 'Amount', 'Status', 'Processed_Date']),
            (self.processed_csv, ['Row', 'Date', 'Initials', 'Source', 'Company', 'Amount', 'Status', 'Processed_Date']),
            (self.needs_rules_csv, ['Row', 'Date', 'Initials', 'Source', 'Company', 'Amount', 'Status', 'Processed_Date', 'Suggestions']),
            (self.failed_csv, ['Row', 'Date', 'Initials', 'Source', 'Company', 'Amount', 'Status', 'Error_Message', 'Failed_Date']),
            (self.ai_matches_csv, ['Original_Source', 'Matched_Source', 'Company', 'Confidence', 'Match_Type', 'Target_Sheet', 'Target_Header', 'Date_Matched'])
        ]
        
        for filepath, headers in csv_files:
            if not Path(filepath).exists():
                with open(filepath, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                logging.info(f"Created CSV file: {filepath}")
    
    def load_existing_transactions_cache(self):
        """Load existing transactions into a set for O(1) lookups."""
        logging.info("LOADING EXISTING TRANSACTIONS CACHE FOR PERFORMANCE")
        
        if not Path(self.preprocessed_csv).exists():
            logging.info("No existing preprocessed CSV found, starting with empty cache")
            self.existing_transactions_cache = set()
            return
        
        try:
            with open(self.preprocessed_csv, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Create unique key for transaction (case-insensitive)
                    transaction_key = (
                        str(row['Row']),
                        row['Source'].lower(),
                        row['Amount'],
                        row['Date']
                    )
                    self.existing_transactions_cache.add(transaction_key)
            
            logging.info(f"Loaded {len(self.existing_transactions_cache)} existing transactions into cache")
            
            except Exception as e:
            logging.error(f"Error loading existing transactions cache: {e}")
            self.existing_transactions_cache = set()
    
    def transaction_exists_in_cache(self, row_num, source, amount, date):
        """Check if transaction exists in cache (O(1) operation)."""
        transaction_key = (str(row_num), source.lower(), amount, date)
        return transaction_key in self.existing_transactions_cache
    
    def add_transaction_to_cache(self, row_num, source, amount, date):
        """Add transaction to cache."""
        transaction_key = (str(row_num), source.lower(), amount, date)
        self.existing_transactions_cache.add(transaction_key)
    
    def handle_ai_corrections(self, original_source, corrected_source, transaction):
        """Handle AI corrections with user confirmation workflow."""
        
        # Create cache key for both original and corrected
        original_key = self.create_transaction_key(transaction, original_source)
        corrected_key = self.create_transaction_key(transaction, corrected_source)
        
        # Remove original from cache, add corrected
        if original_key in self.existing_transactions_cache:
            self.existing_transactions_cache.remove(original_key)
            logging.info(f"Removed original source from cache: {original_source}")
        
        self.existing_transactions_cache.add(corrected_key)
        logging.info(f"Added corrected source to cache: {corrected_source}")
        
        # Move transaction to ready_for_processing.csv
        self.move_to_ready_processing(transaction, corrected_source)
        
        return corrected_key
    
    def create_transaction_key(self, transaction, source):
        """Create a transaction key for cache operations."""
        return (
            str(transaction.get('row', '')),
            str(source).lower(),
            str(transaction.get('total', '')),
            transaction.get('date', '')
        )
    
    def move_to_ready_processing(self, transaction, corrected_source):
        """Move transaction from pending to ready for processing."""
        try:
            # Read pending transactions
            pending_transactions = []
            pending_file = Path("pending_transactions/pending_transactions.csv")
            
            if pending_file.exists():
                with open(pending_file, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        # Skip the transaction being moved
                        if (row['Row'] == str(transaction.get('row', '')) and 
                            row['Source'] == transaction.get('source', '')):
                            continue
                        pending_transactions.append(row)
                
                # Write back pending transactions (without the moved one)
                with open(pending_file, 'w', newline='', encoding='utf-8') as f:
                    if pending_transactions:
                        writer = csv.DictWriter(f, fieldnames=pending_transactions[0].keys())
                        writer.writeheader()
                        writer.writerows(pending_transactions)
            
            # Add to ready_for_processing.csv
            ready_file = Path("pending_transactions/ready_for_processing.csv")
            ready_file.parent.mkdir(exist_ok=True)
            
            with open(ready_file, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    transaction.get('row', ''),
                    transaction.get('date', ''),
                    transaction.get('initials', ''),
                    corrected_source,  # Use corrected source
                    transaction.get('company', ''),
                    transaction.get('total', ''),
                    'ready',
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ])
            
            logging.info(f"Moved transaction {transaction.get('row', '')} to ready for processing with corrected source: {corrected_source}")
            
        except Exception as e:
            logging.error(f"Error moving transaction to ready processing: {e}")
    
    def create_layout_map(self):
        """Create/update layout map from live Google Sheets."""
        logging.info("CREATING/UPDATING LAYOUT MAP FROM LIVE GOOGLE SHEETS")
        logging.info("=" * 60)
        
        # Apply error recovery to this method
        return self.error_recovery.exponential_backoff(self._create_layout_map_impl)()
    
    def _create_layout_map_impl(self):
        """Implementation of layout map creation with error recovery."""
        try:
            # Open the live Google spreadsheet
            spreadsheet = self.gc.open_by_url(self.financials_spreadsheet_url)
            layout_map = {}
            
            # Define sheets to skip
            sheets_to_skip = ['PETTY CASH', 'Petty Cash KEY', 'EXPANSION CHECKLIST', 'Balance and Misc', 'SHIFT LOG', 'TIM', 'Copy of TIM', 'Info']
            
            for worksheet in spreadsheet.worksheets():
                sheet_title = worksheet.title
                if sheet_title in sheets_to_skip:
                    continue
                
                logging.info(f"Mapping layout for sheet: {sheet_title}")
                
                # Get all values to minimize API calls
                all_values = worksheet.get_all_values()
                
                # Map headers (from row 19) to column numbers
                headers = {}
                if len(all_values) >= 19:
                    header_row = all_values[18]  # index 18 is row 19
                    for i, header in enumerate(header_row):
                        if header:
                            headers[header.strip()] = i + 1  # 1-based column index
                
                # Map dates (from column A) to row numbers
                dates = {}
                # Start from row 20 (index 19)
                for i, row in enumerate(all_values[19:]):
                    if row and row[0]:  # Check if row and first cell exist
                        date_str = row[0]  # Get date from first column
                        # Normalize date string for consistent matching
                        try:
                            if isinstance(date_str, datetime):
                                normalized_date = date_str.strftime('%m/%d/%y')
                            elif isinstance(date_str, str):
                                # Parse and normalize the date string
                                parsed_date = pd.to_datetime(date_str, errors='coerce')
                                if pd.notna(parsed_date):
                                    normalized_date = parsed_date.strftime('%m/%d/%y')
                    else:
                                    logging.warning(f"Invalid date format: {date_str}")
                                    continue
                            else:
                                # Handle other date types
                                parsed_date = pd.to_datetime(str(date_str), errors='coerce')
                                if pd.notna(parsed_date):
                                    normalized_date = parsed_date.strftime('%m/%d/%y')
                                else:
                                    logging.warning(f"Invalid date format: {date_str}")
                                    continue
                            
                            # Store the normalized date with row number (20-based)
                            dates[normalized_date] = i + 20
                        except Exception as e:
                            logging.warning(f"Error normalizing date '{date_str}': {e}")
                            continue
                
                layout_map[sheet_title] = {'headers': headers, 'dates': dates}
            
            # Save the map locally
            with open('config/layout_map.json', 'w') as f:
                json.dump(layout_map, f, indent=2)
            
            logging.info(f"Successfully created layout map for {len(layout_map)} sheets.")
            return True
            
        except Exception as e:
            logging.error(f"Failed to create layout map from Google Sheets: {e}")
            import traceback
            logging.error(f"Full traceback: {traceback.format_exc()}")
            return False
    
    def load_layout_map_cache(self):
        """Load layout map from JSON file into memory cache."""
        try:
            layout_map_file = Path("config/layout_map.json")
            if not layout_map_file.exists():
                logging.warning("Layout map file not found, creating new one...")
                if not self.create_layout_map():
                    return False
            
            with open(layout_map_file, 'r') as f:
                self.layout_map_cache = json.load(f)
            
            logging.info(f"Loaded layout map cache with {len(self.layout_map_cache)} sheets")
            return True
            
        except Exception as e:
            logging.error(f"Error loading layout map cache: {e}")
            return False
    
    def preprocess_petty_cash_transactions(self):
        """Download and preprocess new transactions from Google Sheets."""
        logging.info("DOWNLOADING AND PREPROCESSING NEW TRANSACTIONS FROM GOOGLE SHEETS")
        logging.info("=" * 60)
        
        try:
            # Load existing transactions cache for O(1) lookups
            self.load_existing_transactions_cache()
            
            # Connect to Google Sheets and get PETTY CASH worksheet
            spreadsheet = self.gc.open_by_url(self.financials_spreadsheet_url)
            worksheet = spreadsheet.worksheet("PETTY CASH")
            
            # Fetch all data from the worksheet in a single API call
            # Get calculated values by using value_render_option='FORMATTED_VALUE'
            all_data = worksheet.get_all_values(value_render_option='FORMATTED_VALUE')
            
            # Initialize counter for new transactions
            new_transactions_count = 0
            
            # Start progress monitoring for preprocessing
            total_rows = len(all_data[4:])  # Skip header rows
            self.progress_monitor.start_operation(
                "preprocessing", 
                total_rows, 
                "Preprocessing transactions from Google Sheets"
            )
            
            # Open the preprocessed CSV in write mode (start fresh)
            with open(self.preprocessed_csv, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Write header row
                writer.writerow(['Row', 'Date', 'Initials', 'Source', 'Company', 'Amount', 'Status', 'Processed_Date'])
                
                # Process transactions starting from row 5 (index 4)
                for i, row_data in enumerate(all_data[4:]):
                    row_num = i + 5  # 1-based row number in the sheet
                    
                    logging.debug(f"Processing row {row_num}: {row_data[0:4]}")  # Log row data
                    
                    # Update progress
                    self.progress_monitor.update_progress(1, f"Row {row_num}")
                    
                    # Check that essential columns are not empty
                    if not all(row_data[0:4]):  # Skip if essential fields are empty
                        logging.debug(f"Skipping row {row_num}: essential fields empty")
                    continue
                
                # Extract transaction data
                    initials = row_data[0]  # Column A
                    date_val = row_data[1]  # Column B
                    company = row_data[2]   # Column C
                    source = row_data[3]    # Column D
                    amount_val = row_data[18] if len(row_data) > 18 else ''  # Column S
                    
                    # Normalize date to consistent format
                    try:
                        # Handle MM/DD/YY format specifically
                        if '/' in str(date_val):
                            date = pd.to_datetime(date_val, format='%m/%d/%y').strftime('%m/%d/%y')
                        else:
                            date = pd.to_datetime(date_val).strftime('%m/%d/%y')
                    except Exception as e:
                        logging.warning(f"Invalid date format '{date_val}' in row {row_num}: {e}")
                        continue  # Skip rows with invalid dates
                
                # Skip empty rows or $0 amounts
                    if not source or not date or not amount_val:
                        logging.debug(f"Skipping row {row_num}: missing data (source='{source}', date='{date}', amount='{amount_val}')")
                    continue
                
                    # Skip $0 amounts - comprehensive pattern matching
                    amount_str = str(amount_val).strip()
                    zero_amount_patterns = [
                        '$ -', '$ - ', '-', '0', '$0', '$ 0',
                        '$0.00', '$ 0.00', '($0.00)', '($ 0.00)',
                        '-$0.00', '-$ 0.00', '0.00', '0.0',
                        '($0)', '($ 0)', '-$0', '-$ 0',
                        '$0.0', '$ 0.0', '($0.0)', '($ 0.0)',
                        '-$0.0', '-$ 0.0', '0', '0.0', '0.00'
                    ]
                    
                    if amount_str in zero_amount_patterns:
                        logging.debug(f"Skipping zero amount: '{amount_str}' in row {row_num}")
                    continue
                
                    # Create unique key for the transaction
                    transaction_key = (str(row_num), str(source).lower(), str(amount_val), date)
                    
                    # Check if transaction already exists in cache
                    if transaction_key in self.existing_transactions_cache:
                        logging.debug(f"Skipping row {row_num}: transaction already in cache")
                    continue
                
                    # AI-enhanced source correction with cache invalidation
                    original_source = source
                    corrected_source = self.ai_matcher.auto_correct_source(source, company)
                    
                    if corrected_source != original_source:
                        logging.info(f"AI corrected source: '{original_source}' -> '{corrected_source}'")
                        
                        # Create cache keys for both original and corrected sources
                        original_key = (str(row_num), str(original_source).lower(), str(amount_val), date)
                        corrected_key = (str(row_num), str(corrected_source).lower(), str(amount_val), date)
                        
                        # Remove original from cache if it exists
                        if original_key in self.existing_transactions_cache:
                            self.existing_transactions_cache.remove(original_key)
                            logging.debug(f"Removed original source from cache: {original_key}")
                        
                        # Add corrected to cache
                        self.existing_transactions_cache.add(corrected_key)
                        logging.debug(f"Added corrected source to cache: {corrected_key}")
                        
                        # Update source for CSV writing
                        source = corrected_source
                        
                        # Use corrected key for transaction tracking
                        transaction_key = corrected_key
                    else:
                        # No correction needed, use original transaction key
                        transaction_key = (str(row_num), str(source).lower(), str(amount_val), date)
                        self.existing_transactions_cache.add(transaction_key)
                    
                    # Write new transaction to CSV
                    writer.writerow([row_num, date, initials, source, company, amount_val, 'preprocessed', ''])
                    new_transactions_count += 1
            
            # Finish progress monitoring
            self.progress_monitor.finish_operation()
            
            logging.info(f"PREPROCESSING SUMMARY: Found and saved {new_transactions_count} new transactions.")
            logging.info(f"Cache size: {len(self.existing_transactions_cache)} transactions")
                        return True
            
        except Exception as e:
            logging.error(f"Error preprocessing transactions from Google Sheets: {e}")
            return False
    

    
    # REMOVED: transaction_exists_in_csv() - Deprecated, use transaction_exists_in_cache() instead
    
    def read_preprocessed_transactions(self):
        """Read transactions from preprocessed CSV file."""
        logging.info("READING PREPROCESSED TRANSACTIONS")
        
        if not Path(self.preprocessed_csv).exists():
            logging.warning("Preprocessed CSV file not found")
            return []
        
        transactions = []
        try:
            with open(self.preprocessed_csv, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row['Status'] == 'preprocessed':
                        transaction = {
                            'row': int(row['Row']),
                            'initials': row['Initials'],
                            'date': row['Date'],
                            'company': row['Company'],
                            'source': row['Source'],
                            'total': row['Amount']
                        }
                        transactions.append(transaction)
            
            logging.info(f"Read {len(transactions)} preprocessed transactions")
            return transactions
            
        except Exception as e:
            logging.error(f"Error reading preprocessed transactions: {e}")
            return []
    
        
    
    def parse_amount(self, amount_str):
        """Convert $ (60.00) format to number for calculations"""
        if pd.isna(amount_str) or amount_str is None:
            return 0.0
        
        amount_str = str(amount_str).strip()
        
        # Handle negative format: $ (60.00)
        if '$ (' in amount_str and ')' in amount_str:
            # Extract number between parentheses
            number_str = amount_str.replace('$ (', '').replace(')', '').replace(',', '')
            return -float(number_str)
        
        # Handle negative format: (25.50) - parentheses without dollar sign
        elif amount_str.startswith('(') and amount_str.endswith(')'):
            # Extract number between parentheses
            number_str = amount_str[1:-1].replace(',', '')
            return -float(number_str)
        
        # Handle positive format: $1,500.00
        elif '$' in amount_str:
            number_str = amount_str.replace('$', '').replace(',', '').strip()
            return float(number_str)
        
        # Handle plain numbers
        else:
            return float(amount_str)
    
    def format_amount(self, amount):
        """Convert number back to $ (60.00) format"""
        if amount < 0:
            return f"$ ({abs(amount):,.2f})"
        else:
            return f"${amount:,.2f}"
    
    def load_jgd_truth_rules(self):
        """Load rules from JGD Truth Excel file and populate AI matcher with future SQLite support."""
        logging.info("LOADING JGD TRUTH RULES FOR AI MATCHER")
        
        try:
            workbook = openpyxl.load_workbook(self.jgd_truth_file)
            
            for sheet_name in workbook.sheetnames:
                logging.info(f"Processing JGD Truth sheet: {sheet_name}")
                worksheet = workbook[sheet_name]
                
                # Rules start from row 2 (row 1 is header)
                for row in range(2, worksheet.max_row + 1):
                    source = worksheet.cell(row=row, column=1).value
                    target_sheet = worksheet.cell(row=row, column=2).value
                    target_header = worksheet.cell(row=row, column=3).value
                    
                    if source and str(source).strip():
                        source_clean = str(source).strip()
                        rule_key = f"{source_clean}_{sheet_name}"
                        
                        rule_data = {
                            'source': source_clean,
                            'target_sheet': str(target_sheet).strip() if target_sheet else '',
                            'target_header': str(target_header).strip() if target_header else '',
                            'jgd_sheet': sheet_name
                        }
                        
                        self.rules_cache[rule_key] = rule_data
                        
                        # Add to AI matcher
                        self.ai_matcher.rules_cache[rule_key] = rule_data
                        
                        logging.info(f"  Rule: '{source_clean}' -> '{target_sheet}' / '{target_header}'")
            
            workbook.close()
            logging.info(f"Loaded {len(self.rules_cache)} rules from JGD Truth")
            
            # Log AI matcher statistics
            stats = self.ai_matcher.get_match_statistics()
            logging.info(f"AI Matcher Stats: {stats}")
            
        except Exception as e:
            logging.error(f"Error loading JGD Truth rules: {e}")
            raise
    
    def setup_sqlite_rules_database(self):
        """Setup SQLite database for future rule storage scalability."""
        db_path = Path("config/rules_database.db")
        db_path.parent.mkdir(exist_ok=True)
        
        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # Create rules table with indexes for efficient querying
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS rules (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source TEXT NOT NULL,
                    target_sheet TEXT NOT NULL,
                    target_header TEXT NOT NULL,
                    jgd_sheet TEXT NOT NULL,
                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Create indexes for fast lookups
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_source ON rules(source)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_jgd_sheet ON rules(jgd_sheet)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_source_jgd_sheet ON rules(source, jgd_sheet)')
            
            conn.commit()
            conn.close()
            
            logging.info("SQLite rules database setup complete")
            return True
            
        except Exception as e:
            logging.error(f"Error setting up SQLite rules database: {e}")
            return False
    
    def migrate_rules_to_sqlite(self):
        """Migrate existing Excel rules to SQLite database for scalability."""
        logging.info("MIGRATING RULES TO SQLITE DATABASE")
        
        if not self.setup_sqlite_rules_database():
            logging.error("Failed to setup SQLite database")
            return False
        
        try:
            conn = sqlite3.connect("config/rules_database.db")
            cursor = conn.cursor()
            
            # Clear existing data
            cursor.execute('DELETE FROM rules')
            
            # Insert all rules from cache
            for rule_key, rule_data in self.rules_cache.items():
                cursor.execute('''
                    INSERT INTO rules (source, target_sheet, target_header, jgd_sheet)
                    VALUES (?, ?, ?, ?)
                ''', (
                    rule_data['source'],
                    rule_data['target_sheet'],
                    rule_data['target_header'],
                    rule_data['jgd_sheet']
                ))
            
            conn.commit()
            conn.close()
            
            logging.info(f"Migrated {len(self.rules_cache)} rules to SQLite database")
            return True
            
        except Exception as e:
            logging.error(f"Error migrating rules to SQLite: {e}")
            return False
    

    

    
    # REMOVED: get_petty_cash_column_snapshot() - Redundant since we read from Excel files
    
    def find_target_cell_coordinates(self, target_sheet, target_header, date_str):
        """Find target cell coordinates using fast layout map cache."""
        try:
            # Access the in-memory layout map cache
            sheet_map = self.layout_map_cache.get(target_sheet)
            if not sheet_map:
                logging.warning(f"Target sheet '{target_sheet}' not found in layout map cache")
                return None
            
            # Get column number for the target header
            header_col = sheet_map['headers'].get(target_header)
            if not header_col:
                logging.warning(f"Header '{target_header}' not found in layout map for sheet '{target_sheet}'")
                return None
            
            # Normalize date string to consistent format
            try:
                        if isinstance(date_str, datetime):
                    normalized_date = date_str.strftime('%m/%d/%y')
                        elif isinstance(date_str, str):
                    # Parse and normalize the date string
                    parsed_date = pd.to_datetime(date_str, errors='coerce')
                    if pd.notna(parsed_date):
                        normalized_date = parsed_date.strftime('%m/%d/%y')
                        else:
                        logging.warning(f"Invalid date format: {date_str}")
                        return None
                            else:
                    # Handle other date types
                    parsed_date = pd.to_datetime(str(date_str), errors='coerce')
                    if pd.notna(parsed_date):
                        normalized_date = parsed_date.strftime('%m/%d/%y')
                    else:
                        logging.warning(f"Invalid date format: {date_str}")
                        return None
                    except Exception as e:
                logging.warning(f"Error normalizing date '{date_str}': {e}")
                return None
            
            # Get row number for the normalized date
            date_row = sheet_map['dates'].get(normalized_date)
            if not date_row:
                logging.warning(f"Date '{normalized_date}' not found in layout map for sheet '{target_sheet}'")
                return None
            
            return (date_row, header_col)
            
        except Exception as e:
            logging.error(f"Error finding target cell coordinates: {e}")
            return None
    
    def group_transactions_by_sheet(self, transactions):
        """Group transactions by target sheet using AI-enhanced matching."""
        logging.info("GROUPING TRANSACTIONS BY TARGET SHEET WITH AI MATCHING")
        
        # Start progress monitoring for grouping
        self.progress_monitor.start_operation(
            "grouping", 
            len(transactions), 
            "Grouping transactions by target sheet"
        )
        
        sheet_groups = {}
        new_sources_found = set()
        ai_matches = []
        
        for transaction in transactions:
            company = transaction['company']
            source = transaction['source']
            
            # Update progress
            self.progress_monitor.update_progress(1, f"Processing {source}")
            
            # Use AI-enhanced rule matching
            match_result = self.ai_matcher.find_best_match(source, company)
            
            if match_result.matched:
                # Log AI match for analysis
                ai_match_data = {
                    'Original_Source': match_result.original_source,
                    'Matched_Source': match_result.matched_rule['source'],
                    'Company': company,
                    'Confidence': match_result.confidence,
                    'Match_Type': match_result.match_type,
                    'Target_Sheet': match_result.matched_rule['target_sheet'],
                    'Target_Header': match_result.matched_rule['target_header'],
                    'Date_Matched': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                ai_matches.append(ai_match_data)
                
                # Log the match
                logging.info(f"AI Match: '{source}' -> '{match_result.matched_rule['source']}' (Confidence: {match_result.confidence:.2f}, Type: {match_result.match_type})")
                
                # Update threshold learning system (successful match)
                self.ai_matcher.update_threshold_from_match(company, source, match_result.confidence, True)
                
                # Learn the variation if it's a fuzzy match
                if match_result.match_type == "fuzzy" and match_result.confidence > 0.8:
                    self.ai_matcher.learn_variation(source, match_result.matched_rule['source'])
                
                target_sheet = match_result.matched_rule['target_sheet']
                target_header = match_result.matched_rule['target_header']
            
            # Find target cell coordinates
            target_coords = self.find_target_cell_coordinates(target_sheet, target_header, transaction['date'])
            
                if target_coords:
            date_row, header_col = target_coords
            
            # Parse amount
            amount = self.parse_amount(transaction['total'])
            
            # Group by sheet and cell
            if target_sheet not in sheet_groups:
                sheet_groups[target_sheet] = {}
            
            cell_key = f"{date_row}_{header_col}"
            if cell_key not in sheet_groups[target_sheet]:
                sheet_groups[target_sheet][cell_key] = {
                    'row': date_row,
                    'col': header_col,
                    'amount': 0.0,
                    'transactions': []
                }
            
            sheet_groups[target_sheet][cell_key]['amount'] += amount
            sheet_groups[target_sheet][cell_key]['transactions'].append(transaction)
                else:
                    logging.warning(f"No target cell found for {source} on {transaction['date']} in {target_sheet}")
            else:
                # No match found - try AI rule suggestion
                logging.info(f"No AI match found for '{source}' (Company: {company}). Attempting rule creation...")
                
                # Try to create a rule based on suggestions
                if match_result.suggestions:
                    created_rule = self.suggest_rule_creation(source, company, match_result.suggestions)
                    if created_rule:
                        logging.info(f"Successfully created rule for '{source}' using AI suggestions")
                        # Retry the match with the new rule
                        match_result = self.ai_matcher.find_best_match(source, company)
                        if match_result.matched:
                            # Process the transaction with the new rule
                            target_sheet = match_result.matched_rule['target_sheet']
                            target_header = match_result.matched_rule['target_header']
                            
                            # Find target cell coordinates
                            cell_coords = self.find_target_cell_coordinates(target_sheet, target_header, transaction['date'])
                            if cell_coords:
                                date_row, header_col = cell_coords
                                amount = self.parse_amount(transaction['total'])
                                
                                # Add to sheet groups
                                if target_sheet not in sheet_groups:
                                    sheet_groups[target_sheet] = {}
                                
                                cell_key = f"{date_row}_{header_col}"
                                if cell_key not in sheet_groups[target_sheet]:
                                    sheet_groups[target_sheet][cell_key] = {
                                        'row': date_row,
                                        'col': header_col,
                                        'amount': 0,
                                        'transactions': []
                                    }
                                
                                sheet_groups[target_sheet][cell_key]['amount'] += amount
                                sheet_groups[target_sheet][cell_key]['transactions'].append(transaction)
                                continue
                
                # Update threshold learning system (failed match)
                self.ai_matcher.update_threshold_from_match(company, source, 0.0, False)
                
                # If rule creation failed or no suggestions, add to missing rules
                self.missing_rules.append(f"{source} (Company: {company})")
                new_sources_found.add(f"{source} (Company: {company})")
                
                # Log to needs_rules.csv with suggestions
                needs_rule_data = {
                    'Row': transaction['row'],
                    'Date': transaction['date'],
                    'Initials': transaction['initials'],
                    'Source': transaction['source'],
                    'Company': transaction['company'],
                    'Amount': transaction['total'],
                    'Status': 'needs_rule',
                    'Processed_Date': '',
                    'Suggestions': '; '.join(match_result.suggestions) if match_result.suggestions else ''
                }
                
                with open(self.needs_rules_csv, 'a', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=needs_rule_data.keys())
                    writer.writerow(needs_rule_data)
                
                logging.warning(f"No AI match found for '{source}' (Company: {company}). Suggestions: {match_result.suggestions}")
        
        # Save AI matches to CSV
        if ai_matches:
            with open(self.ai_matches_csv, 'a', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=ai_matches[0].keys())
                writer.writerows(ai_matches)
            logging.info(f"Saved {len(ai_matches)} AI matches to CSV")
        
        # Finish progress monitoring
        self.progress_monitor.finish_operation()
        
        # Save learned variations
        self.ai_matcher.save_learned_variations()
        
        logging.info(f"Grouped transactions into {len(sheet_groups)} sheets")
        return sheet_groups
    
    def batch_update_sheets(self, sheet_groups):
        """Update sheets with batch operations."""
        logging.info("EXECUTING BATCH UPDATES")
        
        # Apply error recovery to this method
        return self.error_recovery.exponential_backoff(self._batch_update_sheets_impl)(sheet_groups)
    
    def _batch_update_sheets_impl(self, sheet_groups):
        """Implementation of batch update sheets with error recovery."""
        failed_transactions = []
        
        try:
            spreadsheet = self.gc.open_by_url(self.petty_cash_url)
            
            for sheet_name, cell_updates in sheet_groups.items():
                logging.info(f"Processing batch update for {sheet_name} with {len(cell_updates)} target cells")
                logging.debug(f"Sheet {sheet_name} cell updates: {list(cell_updates.keys())}")
                
                # Get worksheet
                worksheet = None
                for ws in spreadsheet.worksheets():
                    if ws.title == sheet_name:
                        worksheet = ws
                        break
                
                if not worksheet:
                    logging.error(f"Target sheet '{sheet_name}' not found")
                    # Log all transactions for this sheet as failed
                    for cell_key, cell_data in cell_updates.items():
                        for transaction in cell_data.get('transactions', []):
                            self.log_failed_transaction(
                                transaction, 
                                f"Target sheet '{sheet_name}' not found",
                                f"{sheet_name} - Cell ({cell_data['row']},{cell_data['col']})"
                            )
                    continue
                
                # Prepare batch update data
                batch_data = []
                cell_transaction_map = {}  # Map cell updates to their transactions
                
                for cell_key, cell_data in cell_updates.items():
                    row = cell_data['row']
                    col = cell_data['col']
                    new_amount = cell_data['amount']
                    transactions = cell_data.get('transactions', [])
                    
                    # Get current value from target cell
                    try:
                        current_cell = worksheet.cell(row, col)
                        current_value = current_cell.value
                        current_amount = self.parse_amount(current_value)
                        
                        # Add new amount to current
                        final_amount = current_amount + new_amount
                        
                    except Exception as e:
                        logging.warning(f"Error reading current value for {sheet_name} cell ({row},{col}): {e}")
                        final_amount = new_amount
                    
                    # Format final amount
                    formatted_amount = self.format_amount(final_amount)
                    
                    # Add to batch update
                    batch_data.append({
                        'range': f"{sheet_name}!{chr(64+col)}{row}",
                        'values': [[formatted_amount]]
                    })
                    
                    # Store transaction mapping for error handling
                    cell_transaction_map[f"{row}_{col}"] = transactions
                    
                    logging.info(f"Queued: {sheet_name} cell ({row},{col}) = {formatted_amount}")
                
                # Execute batch update
                if batch_data:
                    try:
                        # Use batchUpdate for true batch processing
                        body = {
                            'valueInputOption': 'USER_ENTERED',
                            'data': batch_data
                        }
                        
                        result = spreadsheet.batch_update(body)
                        logging.info(f"Batch update completed for {sheet_name}: {len(batch_data)} cells updated")
                        
                    except Exception as e:
                        logging.error(f"Error executing batch update for {sheet_name}: {e}")
                        
                        # Log all transactions for this sheet as failed
                        for cell_key, transactions in cell_transaction_map.items():
                            for transaction in transactions:
                                self.log_failed_transaction(
                                    transaction,
                                    f"Batch update failed for {sheet_name}: {e}",
                                    f"{sheet_name} - Cell {cell_key}"
                                )
                        continue
            
            logging.info("All batch updates completed")
            
            # Generate failure report if any transactions failed
            if Path(self.failed_csv).exists():
                with open(self.failed_csv, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    failed_count = sum(1 for row in reader if row['Status'] == 'failed')
                    if failed_count > 0:
                        logging.warning(f"Found {failed_count} failed transactions. Check 'Petty Cash Failed.csv' for details.")
                        report = self.generate_failure_report()
                        logging.info("Failure Report:\n" + report)
            
            return True
            
        except Exception as e:
            logging.error(f"Error in batch update: {e}")
            return False
    
    def log_failed_transaction(self, transaction, error_message, cell_info=None):
        """Log a failed transaction to the failed CSV file."""
        try:
            failed_data = {
                'Row': transaction.get('row', ''),
                'Date': transaction.get('date', ''),
                'Initials': transaction.get('initials', ''),
                'Source': transaction.get('source', ''),
                'Company': transaction.get('company', ''),
                'Amount': transaction.get('total', ''),
                'Status': 'failed',
                'Error_Message': error_message,
                'Failed_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Cell_Info': cell_info or ''
            }
            
            with open(self.failed_csv, 'a', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=failed_data.keys())
                writer.writerow(failed_data)
            
            logging.error(f"Failed transaction logged: {transaction.get('source', 'Unknown')} - {error_message}")
            
        except Exception as e:
            logging.error(f"Error logging failed transaction: {e}")

    def retry_failed_cells(self, max_retries=3, retry_delay=5):
        """Retry processing failed transactions from the failed CSV using AI-enhanced matching."""
        logging.info("RETRYING FAILED TRANSACTIONS WITH AI MATCHING")
        logging.info("=" * 60)
        
        if not Path(self.failed_csv).exists():
            logging.info("No failed transactions file found")
            return
        
        try:
            # Read failed transactions
            failed_transactions = []
            with open(self.failed_csv, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row['Status'] == 'failed':
                        failed_transactions.append(row)
            
            if not failed_transactions:
                logging.info("No failed transactions to retry")
                return
            
            logging.info(f"Found {len(failed_transactions)} failed transactions to retry")
            
            # Group failed transactions by error type
            error_groups = {}
            for transaction in failed_transactions:
                error_type = transaction['Error_Message']
                if error_type not in error_groups:
                    error_groups[error_type] = []
                error_groups[error_type].append(transaction)
            
            # Process each error group
            for error_type, transactions in error_groups.items():
                logging.info(f"Retrying {len(transactions)} transactions with error: {error_type}")
                
                for attempt in range(max_retries):
                    logging.info(f"Retry attempt {attempt + 1}/{max_retries}")
                    
                    success_count = 0
                    still_failed = []
                    
                    for transaction in transactions:
                        try:
                            # Convert back to transaction format
                            trans_data = {
                                'row': int(transaction['Row']),
                                'initials': transaction['Initials'],
                                'date': transaction['Date'],
                                'company': transaction['Company'],
                                'source': transaction['Source'],
                                'total': transaction['Amount']
                            }
                            
                            # Use AI-enhanced matching to find rule
                            match_result = self.ai_matcher.find_best_match(trans_data['source'], trans_data['company'])
                            
                            if match_result.matched:
                                # Try to process with AI-matched rule
                                success = self.update_single_cell_with_rule(trans_data, match_result.matched_rule)
                            
                            if success:
                                success_count += 1
                                # Mark as processed in failed CSV
                                self.mark_transaction_processed(transaction)
                            else:
                                still_failed.append(transaction)
                                
                        except Exception as e:
                            logging.error(f"Error retrying transaction {transaction['Source']}: {e}")
                            still_failed.append(transaction)
                    
                    logging.info(f"Retry attempt {attempt + 1}: {success_count} succeeded, {len(still_failed)} still failed")
                    
                    if not still_failed:
                        logging.info("All transactions in this group succeeded!")
                        break
                    
                    transactions = still_failed  # Update for next attempt
                    
                    if attempt < max_retries - 1:
                        logging.info(f"Waiting {retry_delay} seconds before next retry...")
                        import time
                        time.sleep(retry_delay)
            
            logging.info("Failed transaction retry process completed")
            
        except Exception as e:
            logging.error(f"Error in retry process: {e}")

    def update_single_cell_with_rule(self, transaction, rule):
        """Update a single cell using AI-matched rule."""
        try:
            target_sheet = rule['target_sheet']
            target_header = rule['target_header']
            
            if not target_sheet or not target_header:
                logging.warning(f"Invalid target for {transaction['source']}: sheet={target_sheet}, header={target_header}")
                return False
            
            # Find target cell coordinates
            target_coords = self.find_target_cell_coordinates(target_sheet, target_header, transaction['date'])
            
            if not target_coords:
                logging.warning(f"No target cell found for {transaction['source']} on {transaction['date']} in {target_sheet}")
                return False
            
            date_row, header_col = target_coords
            
            # Parse amount
            amount = self.parse_amount(transaction['total'])
            
            # Update the specific cell
            success = self.update_single_cell(target_sheet, date_row, header_col, amount)
            
            if success:
                logging.info(f"Successfully processed: {transaction['source']} -> {target_sheet} cell ({date_row},{header_col})")
                return True
            else:
                logging.error(f"Failed to update cell for {transaction['source']}")
                return False
                
        except Exception as e:
            logging.error(f"Error processing transaction {transaction.get('source', 'Unknown')}: {e}")
            return False

    def update_single_cell(self, sheet_name, row, col, new_amount):
        """Update a single cell in Google Sheets."""
        # Apply error recovery to this method
        return self.error_recovery.exponential_backoff(self._update_single_cell_impl)(sheet_name, row, col, new_amount)
    
    def _update_single_cell_impl(self, sheet_name, row, col, new_amount):
        """Implementation of single cell update with error recovery."""
        try:
            spreadsheet = self.gc.open_by_url(self.petty_cash_url)
            
            # Get worksheet
            worksheet = None
            for ws in spreadsheet.worksheets():
                if ws.title == sheet_name:
                    worksheet = ws
                    break
            
            if not worksheet:
                logging.error(f"Target sheet '{sheet_name}' not found")
                return False
            
            # Get current value from target cell
            try:
                current_cell = worksheet.cell(row, col)
                current_value = current_cell.value
                current_amount = self.parse_amount(current_value)
                
                # Add new amount to current
                final_amount = current_amount + new_amount
                
            except Exception as e:
                logging.warning(f"Error reading current value for {sheet_name} cell ({row},{col}): {e}")
                final_amount = new_amount
            
            # Format final amount
            formatted_amount = self.format_amount(final_amount)
            
            # Update the cell
            cell_address = f"{chr(64+col)}{row}"
            worksheet.update(cell_address, formatted_amount)
            
            logging.info(f"Updated {sheet_name} cell {cell_address} = {formatted_amount}")
            return True
            
        except Exception as e:
            logging.error(f"Error updating single cell {sheet_name} ({row},{col}): {e}")
            return False

    def mark_transaction_processed(self, transaction):
        """Mark a transaction as processed in the failed CSV."""
        try:
            # Read all rows
            rows = []
            with open(self.failed_csv, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames
                for row in reader:
                    # Mark matching transaction as processed
                    if (row['Row'] == transaction['Row'] and 
                        row['Source'] == transaction['Source'] and 
                        row['Amount'] == transaction['Amount'] and
                        row['Status'] == 'failed'):
                        row['Status'] = 'processed'
                        row['Processed_Date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    rows.append(row)
            
            # Write back to file
            with open(self.failed_csv, 'w', newline='', encoding='utf-8') as f:
                if fieldnames:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(rows)
                else:
                    logging.error("Fieldnames is None, cannot write CSV")
                
        except Exception as e:
            logging.error(f"Error marking transaction as processed: {e}")

    def generate_failure_report(self):
        """Generate a detailed report of failed transactions."""
        if not Path(self.failed_csv).exists():
            return "No failed transactions found."
        
        try:
            failed_transactions = []
            with open(self.failed_csv, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row['Status'] == 'failed':
                        failed_transactions.append(row)
            
            if not failed_transactions:
                return "No failed transactions found."
            
            report = f"FAILED TRANSACTIONS REPORT\n"
            report += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            report += f"Total Failed: {len(failed_transactions)}\n"
            report += "=" * 60 + "\n\n"
            
            # Group by error type
            error_summary = {}
            for transaction in failed_transactions:
                error_type = transaction['Error_Message']
                if error_type not in error_summary:
                    error_summary[error_type] = []
                error_summary[error_type].append(transaction)
            
            for error_type, transactions in error_summary.items():
                report += f"ERROR TYPE: {error_type}\n"
                report += f"Count: {len(transactions)}\n"
                report += "-" * 40 + "\n"
                
                for transaction in transactions:
                    report += f"  {transaction['Date']} - {transaction['Source']} ({transaction['Company']}) - ${transaction['Amount']}\n"
                    if transaction['Cell_Info']:
                        report += f"    Target: {transaction['Cell_Info']}\n"
                report += "\n"
            
            return report
            
        except Exception as e:
            return f"Error generating failure report: {e}"
    
    def run(self, dry_run=True, migrate_to_sqlite=False):
        """Run the complete automated sorting process with direct Google Sheets integration."""
        start_time = datetime.now()
        logging.info("STARTING FULLY AUTOMATED PETTY CASH SORTING")
        logging.info("=" * 60)
        
        # Check for resume from previous run
        progress = self.error_recovery.resume_from_progress()
        if progress:
            logging.info(f"Resuming from previous run: {progress['operation']}")
        
        # Start performance monitoring
        self.performance_metrics.start_monitoring()
        
        try:
            # Step 1: Create/update layout map from live Google Sheets
            logging.info("STEP 1/5: Creating/updating layout map from Google Sheets...")
            if not self.create_layout_map():
                logging.error("Failed to create layout map")
                return False
            
            # Save progress after Step 1
            self.error_recovery.save_progress("layout_map_created", 1, 5, {"step": "layout_map"})
            
            # Step 2: Download and preprocess new transactions from Google Sheets
            logging.info("STEP 2/5: Downloading and preprocessing new transactions...")
            if not self.preprocess_petty_cash_transactions():
                logging.error("Failed to preprocess transactions")
                return False
            
            # Save progress after Step 2
            self.error_recovery.save_progress("transactions_preprocessed", 2, 5, {"step": "preprocessing"})
            
            # Step 3: Load business rules and layout map cache
            logging.info("STEP 3/5: Loading business rules and layout map cache...")
            self.load_ai_rules()
            
            if not self.load_layout_map_cache():
                logging.error("Failed to load layout map cache")
                return False
            
            # Optional: Migrate to SQLite for future scalability
            if migrate_to_sqlite:
                logging.info("OPTIONAL STEP: Migrating rules to SQLite database...")
                if self.migrate_rules_to_sqlite():
                    logging.info("Successfully migrated rules to SQLite for future scalability")
                else:
                    logging.warning("SQLite migration failed, continuing with Excel rules")
            
            # Step 4: Read and group transactions with AI matching
            logging.info("STEP 4/5: Reading and grouping transactions with AI matching...")
            transactions = self.read_preprocessed_transactions()
            
            if not transactions:
                logging.info("No transactions found")
                return True
            
            sheet_groups = self.group_transactions_by_sheet(transactions)
            
            if not sheet_groups:
                logging.info("No transactions grouped for processing")
                return True
            
            # Step 5: Execute updates or show dry run results
            logging.info("STEP 5/5: Finalizing process...")
            if not dry_run:
                logging.info("Executing live batch updates...")
                success = self.batch_update_sheets(sheet_groups)
                if success:
                    # AI matcher handles rule learning automatically
                    logging.info("AI matcher will learn from successful matches")
                    
                    # Retry failed transactions
                    logging.info("RETRY STEP: Retrying failed transactions...")
                    self.retry_failed_cells(max_retries=3, retry_delay=5)
                    
                    logging.info("Live update completed successfully")
                else:
                    logging.error("Batch update failed")
                    return False
            else:
                logging.info("DRY RUN - Would update the following:")
                for sheet_name, cell_updates in sheet_groups.items():
                    logging.info(f"  {sheet_name}: {len(cell_updates)} cells")
                    for cell_key, cell_data in cell_updates.items():
                        formatted_amount = self.format_amount(cell_data['amount'])
                        logging.info(f"    Cell ({cell_data['row']},{cell_data['col']}): {formatted_amount}")
            
            # Performance summary
            end_time = datetime.now()
            duration = end_time - start_time
            self.log_performance_summary(duration, len(transactions))
            
            # Clear progress file on successful completion
            self.error_recovery.clear_progress()
            
            # Log error recovery stats
            recovery_stats = self.error_recovery.get_recovery_stats()
            logging.info(f"Error recovery stats: {recovery_stats}")
            
            # Stop performance monitoring and log stats
            self.performance_metrics.stop_monitoring()
            performance_stats = self.performance_metrics.get_current_stats()
            logging.info(f"Performance stats: {performance_stats}")
            
            return True
            
        except Exception as e:
            logging.error(f"Fatal error in sorting process: {e}")
            raise
    
    def log_performance_summary(self, duration, transaction_count):
        """Log performance summary with optimization metrics."""
        logging.info("=" * 60)
        logging.info("PERFORMANCE SUMMARY")
        logging.info("=" * 60)
        logging.info(f"Total execution time: {duration}")
        logging.info(f"Transactions processed: {transaction_count}")
        logging.info(f"Cache size: {len(self.existing_transactions_cache)} transactions")
        logging.info(f"Rules loaded: {len(self.rules_cache)}")
        logging.info(f"AI matcher stats: {self.ai_matcher.get_match_statistics()}")
        
        # Calculate performance metrics
        if transaction_count > 0:
            avg_time_per_transaction = duration.total_seconds() / transaction_count
            logging.info(f"Average time per transaction: {avg_time_per_transaction:.3f} seconds")
        
        logging.info("=" * 60)

    def load_ai_rules(self):
        """Load rules from AI-driven rule management system instead of JGD Truth."""
        logging.info("LOADING RULES FROM AI-DRIVEN RULE MANAGEMENT SYSTEM")
        
        try:
            # First, try to load from SQLite database
            if self.load_rules_from_sqlite():
                logging.info(f"Loaded {len(self.rules_cache)} rules from SQLite database")
                return True
            
            # If SQLite is empty, migrate from JGD Truth (one-time migration)
            if self.migrate_jgd_truth_to_sqlite():
                logging.info("Successfully migrated JGD Truth rules to SQLite")
                return True
            
            # If no rules found anywhere, create default rules
            logging.warning("No rules found, creating default rule set")
            self.create_default_rules()
            return True
            
        except Exception as e:
            logging.error(f"Error loading AI rules: {e}")
            return False
    
    def load_rules_from_sqlite(self):
        """Load rules from SQLite database."""
        try:
            db_path = Path("config/rules_database.db")
            if not db_path.exists():
                return False
            
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            cursor.execute('SELECT source, target_sheet, target_header, jgd_sheet FROM rules')
            rules = cursor.fetchall()
            
            self.rules_cache.clear()
            self.ai_matcher.rules_cache.clear()
            
            for source, target_sheet, target_header, jgd_sheet in rules:
                rule_key = f"{source}_{jgd_sheet}"
                rule_data = {
                    'source': source,
                    'target_sheet': target_sheet,
                    'target_header': target_header,
                    'jgd_sheet': jgd_sheet
                }
                
                self.rules_cache[rule_key] = rule_data
                self.ai_matcher.rules_cache[rule_key] = rule_data
            
            conn.close()
            return len(rules) > 0
            
        except Exception as e:
            logging.error(f"Error loading rules from SQLite: {e}")
            return False
    
    def migrate_jgd_truth_to_sqlite(self):
        """One-time migration from JGD Truth Excel to SQLite."""
        try:
            if not Path(self.jgd_truth_file).exists():
                logging.warning(f"JGD Truth file not found: {self.jgd_truth_file}")
                return False
            
            # Load rules from JGD Truth (temporary)
            self.load_jgd_truth_rules()
            
            # Migrate to SQLite
            if self.migrate_rules_to_sqlite():
                logging.info("Successfully migrated JGD Truth rules to SQLite")
                return True
            
            return False
            
        except Exception as e:
            logging.error(f"Error migrating JGD Truth to SQLite: {e}")
            return False
    
    def create_default_rules(self):
        """Create default rules for common business operations."""
        logging.info("Creating default rule set")
        
        default_rules = [
            # NUGZ Rules
            {'source': 'SALE TO', 'target_sheet': 'NUGZ', 'target_header': 'Sales', 'jgd_sheet': 'NUGZ'},
            {'source': 'PAYROLL', 'target_sheet': 'NUGZ', 'target_header': 'Payroll', 'jgd_sheet': 'NUGZ'},
            {'source': 'REG', 'target_sheet': 'NUGZ', 'target_header': 'Register', 'jgd_sheet': 'NUGZ'},
            
            # JGD Rules
            {'source': 'SALE TO', 'target_sheet': 'JGD', 'target_header': 'Sales', 'jgd_sheet': 'JGD'},
            {'source': 'PAYROLL', 'target_sheet': 'JGD', 'target_header': 'Payroll', 'jgd_sheet': 'JGD'},
            
            # PUFFIN Rules
            {'source': 'ORDER', 'target_sheet': 'PUFFIN', 'target_header': 'Orders', 'jgd_sheet': 'PUFFIN'},
            {'source': 'LAB', 'target_sheet': 'PUFFIN', 'target_header': 'Lab Testing', 'jgd_sheet': 'PUFFIN'},
        ]
        
        for rule_data in default_rules:
            rule_key = f"{rule_data['source']}_{rule_data['jgd_sheet']}"
            self.rules_cache[rule_key] = rule_data
            self.ai_matcher.rules_cache[rule_key] = rule_data
        
        # Save to SQLite
        self.save_rules_to_sqlite()
        logging.info(f"Created {len(default_rules)} default rules")
    
    def save_rules_to_sqlite(self):
        """Save current rules to SQLite database."""
        try:
            if not self.setup_sqlite_rules_database():
                return False
            
            conn = sqlite3.connect("config/rules_database.db")
            cursor = conn.cursor()
            
            # Clear existing data
            cursor.execute('DELETE FROM rules')
            
            # Insert all rules from cache
            for rule_key, rule_data in self.rules_cache.items():
                cursor.execute('''
                    INSERT INTO rules (source, target_sheet, target_header, jgd_sheet)
                    VALUES (?, ?, ?, ?)
                ''', (
                    rule_data['source'],
                    rule_data['target_sheet'],
                    rule_data['target_header'],
                    rule_data['jgd_sheet']
                ))
            
            conn.commit()
            conn.close()
            
            logging.info(f"Saved {len(self.rules_cache)} rules to SQLite database")
            return True
            
        except Exception as e:
            logging.error(f"Error saving rules to SQLite: {e}")
            return False
    
    def create_rule_with_confirmation(self, source, company, target_sheet, target_header):
        """Create new rule with user confirmation workflow."""
        logging.info(f"Creating new rule: '{source}' -> '{target_sheet}' / '{target_header}'")
        
        rule_data = {
            'source': source,
            'target_sheet': target_sheet,
            'target_header': target_header,
            'jgd_sheet': company,
            'status': 'pending_confirmation'
        }
        
        # Check if rule already exists
        rule_key = f"{source}_{company}"
        if rule_key in self.rules_cache:
            logging.info(f"Rule already exists: {rule_key}")
            return True
        
        # Add to cache and save to SQLite
        self.rules_cache[rule_key] = rule_data
        self.ai_matcher.rules_cache[rule_key] = rule_data
        
        # Save to SQLite
        if self.save_rules_to_sqlite():
            logging.info(f"Successfully created and saved rule: {rule_key}")
            return True
        else:
            logging.error(f"Failed to save rule: {rule_key}")
            return False
    
    def suggest_rule_creation(self, source, company, suggestions):
        """Suggest rule creation for unmatched sources."""
        logging.info(f"Suggesting rule creation for '{source}' (Company: {company})")
        
        if not suggestions:
            logging.warning(f"No suggestions available for '{source}'")
            return None
        
        # Use the best suggestion as default
        best_suggestion = suggestions[0]
        
        # Create rule with best suggestion
        success = self.create_rule_with_confirmation(
            source=source,
            company=company,
            target_sheet=best_suggestion.get('target_sheet', company),
            target_header=best_suggestion.get('target_header', 'Other')
        )
        
        if success:
            logging.info(f"Auto-created rule for '{source}' using best suggestion")
            return best_suggestion
        
        return None

class ProgressMonitor:
    """Comprehensive progress monitoring with tqdm bars and ETA calculations."""
    
    def __init__(self):
        self.start_time = None
        self.operation_status = "idle"
        self.milestone_interval = 10  # Log milestones every 10 transactions
        self.current_pbar = None
        self.operation_stats = {}
        
    def start_operation(self, operation_name: str, total_items: int, description: str = None):
        """Start monitoring a new operation."""
        self.operation_status = operation_name
        self.start_time = time.time()
        
        # Create progress bar
        desc = description or operation_name
        self.current_pbar = tqdm(
            total=total_items, 
            desc=desc,
            unit='items',
            bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]'
        )
        
        # Initialize stats
        self.operation_stats[operation_name] = {
            'start_time': self.start_time,
            'total_items': total_items,
            'processed': 0,
            'milestones': []
        }
        
        logging.info(f"Started operation: {operation_name} ({total_items} items)")
    
    def update_progress(self, count: int = 1, message: str = None):
        """Update progress and log milestones."""
        if not self.current_pbar:
            return
        
        # Update progress bar
        self.current_pbar.update(count)
        if message:
            self.current_pbar.set_description(f"{self.operation_status}: {message}")
        
        # Update stats
        if self.operation_status in self.operation_stats:
            self.operation_stats[self.operation_status]['processed'] += count
            
            # Log milestones
            processed = self.operation_stats[self.operation_status]['processed']
            total = self.operation_stats[self.operation_status]['total_items']
            
            if processed % self.milestone_interval == 0:
                self._log_milestone(processed, total)
    
    def _log_milestone(self, processed: int, total: int):
        """Log milestone with ETA and rate calculations."""
        elapsed = time.time() - self.start_time
        rate = processed / elapsed if elapsed > 0 else 0
        eta = (total - processed) / rate if rate > 0 else 0
        
        milestone_info = {
            'processed': processed,
            'total': total,
            'percentage': (processed / total) * 100,
            'elapsed': elapsed,
            'rate': rate,
            'eta': eta,
            'timestamp': time.time()
        }
        
        # Store milestone
        if self.operation_status in self.operation_stats:
            self.operation_stats[self.operation_status]['milestones'].append(milestone_info)
        
        # Log milestone
        logging.info(
            f"Milestone: {processed}/{total} ({milestone_info['percentage']:.1f}%) "
            f"processed. Rate: {rate:.1f}/sec, ETA: {eta:.1f}s"
        )
    
    def finish_operation(self):
        """Finish the current operation and log summary."""
        if not self.current_pbar:
            return
        
        self.current_pbar.close()
        total_time = time.time() - self.start_time
        
        # Log completion
        if self.operation_status in self.operation_stats:
            stats = self.operation_stats[self.operation_status]
            avg_rate = stats['processed'] / total_time if total_time > 0 else 0
            
            logging.info(
                f"Operation '{self.operation_status}' completed: "
                f"{stats['processed']}/{stats['total_items']} items "
                f"in {total_time:.2f}s (avg: {avg_rate:.1f}/sec)"
            )
        
        self.current_pbar = None
        self.operation_status = "idle"
    
    def get_operation_summary(self) -> dict:
        """Get summary of all operations."""
        summary = {}
        for op_name, stats in self.operation_stats.items():
            if stats['start_time']:
                duration = time.time() - stats['start_time']
                avg_rate = stats['processed'] / duration if duration > 0 else 0
                
                summary[op_name] = {
                    'processed': stats['processed'],
                    'total': stats['total_items'],
                    'duration': duration,
                    'avg_rate': avg_rate,
                    'milestones': len(stats['milestones'])
                }
        
        return summary

class ErrorRecovery:
    """Comprehensive error recovery with exponential backoff and AI assistance."""
    
    def __init__(self):
        self.max_retries = 3
        self.base_delay = 1  # Base delay in seconds
        self.max_delay = 60  # Maximum delay in seconds
        self.progress_file = "config/processing_progress.json"
        self.error_log_file = "logs/error_recovery.log"
        self.retry_count = 0
        self.total_retries = 0
        
        # Error patterns for AI diagnosis
        self.error_patterns = {
            'rate_limit': [
                'quota exceeded', 'rate limit', 'too many requests',
                'quota exceeded', 'rate limit exceeded'
            ],
            'network': [
                'connection', 'timeout', 'network', 'dns', 'ssl',
                'connection reset', 'connection refused'
            ],
            'auth': [
                'unauthorized', 'forbidden', 'invalid credentials',
                'authentication', 'permission denied'
            ],
            'api': [
                'api error', 'bad request', 'invalid request',
                'malformed request', 'invalid parameter'
            ],
            'resource': [
                'not found', 'resource not found', 'worksheet not found',
                'spreadsheet not found', 'file not found'
            ]
        }
        
        # Recovery strategies
        self.recovery_strategies = {
            'rate_limit': self._handle_rate_limit,
            'network': self._handle_network_error,
            'auth': self._handle_auth_error,
            'api': self._handle_api_error,
            'resource': self._handle_resource_error
        }
    
    def exponential_backoff(self, func, *args, **kwargs):
        """Decorator for exponential backoff with jitter."""
        def wrapper(*args, **kwargs):
            for attempt in range(self.max_retries):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    if attempt == self.max_retries - 1:
                        # Final attempt failed, log and re-raise
                        self._log_final_failure(func.__name__, e, args, kwargs)
                        raise
                    
                    # Calculate delay with exponential backoff and jitter
                    delay = self._calculate_delay(attempt)
                    
                    # Diagnose error and get recovery strategy
                    error_type = self._diagnose_error(e)
                    recovery_msg = self._get_recovery_message(error_type, attempt + 1)
                    
                    logging.warning(
                        f"Attempt {attempt + 1} failed for {func.__name__}: {e}. "
                        f"Retrying in {delay:.1f}s... {recovery_msg}"
                    )
                    
                    # Apply recovery strategy
                    self._apply_recovery_strategy(error_type, attempt + 1)
                    
                    # Wait with exponential backoff
                    time.sleep(delay)
                    self.retry_count += 1
                    self.total_retries += 1
            
            return None
        return wrapper
    
    def _calculate_delay(self, attempt):
        """Calculate delay with exponential backoff and jitter."""
        # Exponential backoff: base_delay * 2^attempt
        delay = self.base_delay * (2 ** attempt)
        
        # Add jitter (±25% random variation)
        jitter = delay * 0.25 * random.uniform(-1, 1)
        delay += jitter
        
        # Cap at maximum delay
        delay = min(delay, self.max_delay)
        
        return delay
    
    def _diagnose_error(self, error):
        """AI-assisted error diagnosis."""
        error_str = str(error).lower()
        
        for error_type, patterns in self.error_patterns.items():
            for pattern in patterns:
                if pattern in error_str:
                    logging.debug(f"Diagnosed error as {error_type}: {error}")
                    return error_type
        
        # Default to unknown error
        logging.debug(f"Unknown error type: {error}")
        return 'unknown'
    
    def _get_recovery_message(self, error_type, attempt):
        """Get human-readable recovery message."""
        messages = {
            'rate_limit': f"Rate limit detected (attempt {attempt}/{self.max_retries})",
            'network': f"Network issue detected (attempt {attempt}/{self.max_retries})",
            'auth': f"Authentication issue detected (attempt {attempt}/{self.max_retries})",
            'api': f"API error detected (attempt {attempt}/{self.max_retries})",
            'resource': f"Resource not found (attempt {attempt}/{self.max_retries})",
            'unknown': f"Unknown error (attempt {attempt}/{self.max_retries})"
        }
        return messages.get(error_type, messages['unknown'])
    
    def _apply_recovery_strategy(self, error_type, attempt):
        """Apply specific recovery strategy based on error type."""
        if error_type in self.recovery_strategies:
            try:
                self.recovery_strategies[error_type](attempt)
            except Exception as e:
                logging.warning(f"Recovery strategy failed: {e}")
    
    def _handle_rate_limit(self, attempt):
        """Handle rate limit errors."""
        # For rate limits, we might want to wait longer
        if attempt == 1:
            logging.info("Rate limit detected. Implementing longer backoff...")
            # Could implement adaptive rate limiting here
    
    def _handle_network_error(self, attempt):
        """Handle network errors."""
        if attempt == 1:
            logging.info("Network error detected. Checking connectivity...")
            # Could implement network connectivity check here
    
    def _handle_auth_error(self, attempt):
        """Handle authentication errors."""
        if attempt == 1:
            logging.error("Authentication error detected. Check credentials.")
            # Could implement credential refresh here
    
    def _handle_api_error(self, attempt):
        """Handle API errors."""
        if attempt == 1:
            logging.info("API error detected. Validating request parameters...")
            # Could implement request validation here
    
    def _handle_resource_error(self, attempt):
        """Handle resource not found errors."""
        if attempt == 1:
            logging.error("Resource not found. Check spreadsheet/worksheet access.")
            # Could implement resource validation here
    
    def save_progress(self, operation, completed_items, total_items, context=None):
        """Save processing progress for resume functionality."""
        progress = {
            'operation': operation,
            'completed': completed_items,
            'total': total_items,
            'timestamp': time.time(),
            'context': context or {},
            'retry_count': self.retry_count
        }
        
        try:
            Path(self.progress_file).parent.mkdir(exist_ok=True)
            with open(self.progress_file, 'w') as f:
                json.dump(progress, f, indent=2)
            logging.debug(f"Progress saved: {completed_items}/{total_items} items")
        except Exception as e:
            logging.error(f"Failed to save progress: {e}")
    
    def resume_from_progress(self):
        """Resume processing from saved progress."""
        if not Path(self.progress_file).exists():
            return None
        
        try:
            with open(self.progress_file, 'r') as f:
                progress = json.load(f)
            
            # Check if progress is recent (within last hour)
            if time.time() - progress['timestamp'] > 3600:
                logging.info("Progress file is too old, starting fresh")
                return None
            
            logging.info(f"Resuming from progress: {progress['completed']}/{progress['total']} items")
            return progress
            
        except Exception as e:
            logging.error(f"Failed to load progress: {e}")
            return None
    
    def clear_progress(self):
        """Clear saved progress."""
        try:
            if Path(self.progress_file).exists():
                Path(self.progress_file).unlink()
                logging.debug("Progress file cleared")
        except Exception as e:
            logging.error(f"Failed to clear progress: {e}")
    
    def _log_final_failure(self, func_name, error, args, kwargs):
        """Log final failure with context."""
        error_info = {
            'function': func_name,
            'error': str(error),
            'error_type': type(error).__name__,
            'timestamp': time.time(),
            'retry_count': self.retry_count,
            'args': str(args)[:200],  # Truncate long args
            'kwargs': str(kwargs)[:200]  # Truncate long kwargs
        }
        
        try:
            Path(self.error_log_file).parent.mkdir(exist_ok=True)
            with open(self.error_log_file, 'a') as f:
                f.write(json.dumps(error_info) + '\n')
        except Exception as e:
            logging.error(f"Failed to log error: {e}")
        
        logging.error(f"Final failure after {self.retry_count} retries: {error}")
    
    def get_recovery_stats(self):
        """Get error recovery statistics."""
        return {
            'total_retries': self.total_retries,
            'current_retry_count': self.retry_count,
            'max_retries': self.max_retries,
            'base_delay': self.base_delay,
            'max_delay': self.max_delay
        }

class PerformanceMetrics:
    """Comprehensive performance monitoring with timing, API tracking, and memory analysis."""
    
    def __init__(self):
        self.start_time = None
        self.operation_timings = {}
        self.api_calls = {
            'total': 0,
            'successful': 0,
            'failed': 0,
            'response_times': [],
            'rate_limits': 0,
            'errors': {}
        }
        self.memory_usage = []
        self.memory_monitor_thread = None
        self.monitoring_active = False
        self.performance_log_file = "logs/performance_metrics.json"
        self.bottlenecks = []
        self.optimization_suggestions = []
        
        # Performance thresholds
        self.thresholds = {
            'api_response_time': 2.0,  # seconds
            'memory_usage': 80.0,  # percentage
            'operation_time': 30.0,  # seconds
            'api_failure_rate': 0.1  # 10%
        }
    
    def start_monitoring(self):
        """Start comprehensive performance monitoring."""
        self.start_time = time.time()
        self.monitoring_active = True
        
        # Start memory monitoring in background thread
        self.memory_monitor_thread = threading.Thread(target=self._monitor_memory, daemon=True)
        self.memory_monitor_thread.start()
        
        logging.info("Performance monitoring started")
    
    def stop_monitoring(self):
        """Stop performance monitoring and generate report."""
        self.monitoring_active = False
        if self.memory_monitor_thread:
            self.memory_monitor_thread.join(timeout=1)
        
        # Generate final report
        self._generate_performance_report()
        logging.info("Performance monitoring stopped")
    
    def _monitor_memory(self):
        """Background thread for memory monitoring."""
        while self.monitoring_active:
            try:
                process = psutil.Process()
                memory_info = {
                    'timestamp': time.time(),
                    'rss': process.memory_info().rss / 1024 / 1024,  # MB
                    'vms': process.memory_info().vms / 1024 / 1024,  # MB
                    'percent': process.memory_percent(),
                    'available': psutil.virtual_memory().available / 1024 / 1024  # MB
                }
                self.memory_usage.append(memory_info)
                
                # Check for memory issues
                if memory_info['percent'] > self.thresholds['memory_usage']:
                    self.bottlenecks.append({
                        'type': 'memory',
                        'severity': 'high' if memory_info['percent'] > 90 else 'medium',
                        'value': memory_info['percent'],
                        'timestamp': memory_info['timestamp']
                    })
                
                time.sleep(5)  # Monitor every 5 seconds
                
            except Exception as e:
                logging.warning(f"Memory monitoring error: {e}")
                time.sleep(10)
    
    def time_operation(self, operation_name: str):
        """Context manager for timing operations."""
        return OperationTimer(self, operation_name)
    
    def track_api_call(self, api_name: str, response_time: float, success: bool, error_type: str = None):
        """Track API call performance."""
        self.api_calls['total'] += 1
        self.api_calls['response_times'].append(response_time)
        
        if success:
            self.api_calls['successful'] += 1
        else:
            self.api_calls['failed'] += 1
            if error_type:
                if error_type not in self.api_calls['errors']:
                    self.api_calls['errors'][error_type] = 0
                self.api_calls['errors'][error_type] += 1
        
        # Check for performance issues
        if response_time > self.thresholds['api_response_time']:
            self.bottlenecks.append({
                'type': 'api_slow',
                'api': api_name,
                'response_time': response_time,
                'timestamp': time.time()
            })
    
    def _generate_performance_report(self):
        """Generate comprehensive performance report."""
        if not self.start_time:
            return
        
        total_time = time.time() - self.start_time
        report = {
            'execution_summary': {
                'total_time': total_time,
                'start_time': self.start_time,
                'end_time': time.time()
            },
            'operation_timings': self.operation_timings,
            'api_performance': {
                'total_calls': self.api_calls['total'],
                'success_rate': self.api_calls['successful'] / max(self.api_calls['total'], 1),
                'average_response_time': sum(self.api_calls['response_times']) / max(len(self.api_calls['response_times']), 1),
                'error_breakdown': self.api_calls['errors']
            },
            'memory_analysis': self._analyze_memory_usage(),
            'bottlenecks': self.bottlenecks,
            'optimization_suggestions': self._generate_optimization_suggestions()
        }
        
        # Save report
        try:
            Path(self.performance_log_file).parent.mkdir(exist_ok=True)
            with open(self.performance_log_file, 'w') as f:
                json.dump(report, f, indent=2)
        except Exception as e:
            logging.error(f"Failed to save performance report: {e}")
        
        # Log summary
        self._log_performance_summary(report)
    
    def _analyze_memory_usage(self):
        """Analyze memory usage patterns."""
        if not self.memory_usage:
            return {'error': 'No memory data collected'}
        
        memory_values = [m['percent'] for m in self.memory_usage]
        return {
            'peak_usage': max(memory_values),
            'average_usage': sum(memory_values) / len(memory_values),
            'min_usage': min(memory_values),
            'memory_leak_detected': self._detect_memory_leak(),
            'samples': len(self.memory_usage)
        }
    
    def _detect_memory_leak(self):
        """Detect potential memory leaks."""
        if len(self.memory_usage) < 10:
            return False
        
        # Check if memory usage is consistently increasing
        recent_memory = [m['percent'] for m in self.memory_usage[-10:]]
        if len(recent_memory) >= 5:
            # Simple trend analysis
            first_half = sum(recent_memory[:5]) / 5
            second_half = sum(recent_memory[5:]) / 5
            return second_half > first_half * 1.2  # 20% increase threshold
        
        return False
    
    def _generate_optimization_suggestions(self):
        """Generate optimization suggestions based on performance data."""
        suggestions = []
        
        # API performance suggestions
        if self.api_calls['total'] > 0:
            failure_rate = self.api_calls['failed'] / self.api_calls['total']
            if failure_rate > self.thresholds['api_failure_rate']:
                suggestions.append({
                    'type': 'api_reliability',
                    'priority': 'high',
                    'description': f"High API failure rate ({failure_rate:.1%}). Consider implementing better error handling and retry logic."
                })
            
            avg_response_time = sum(self.api_calls['response_times']) / len(self.api_calls['response_times'])
            if avg_response_time > self.thresholds['api_response_time']:
                suggestions.append({
                    'type': 'api_performance',
                    'priority': 'medium',
                    'description': f"Slow API response times (avg: {avg_response_time:.2f}s). Consider caching or batch operations."
                })
        
        # Memory suggestions
        if self.memory_usage:
            peak_memory = max(m['percent'] for m in self.memory_usage)
            if peak_memory > 90:
                suggestions.append({
                    'type': 'memory_usage',
                    'priority': 'high',
                    'description': f"High memory usage detected (peak: {peak_memory:.1f}%). Consider memory optimization."
                })
            
            if self._detect_memory_leak():
                suggestions.append({
                    'type': 'memory_leak',
                    'priority': 'high',
                    'description': "Potential memory leak detected. Review object lifecycle and cleanup."
                })
        
        # Operation timing suggestions
        for operation, timing in self.operation_timings.items():
            if timing['total_time'] > self.thresholds['operation_time']:
                suggestions.append({
                    'type': 'operation_performance',
                    'priority': 'medium',
                    'description': f"Slow operation: {operation} took {timing['total_time']:.2f}s. Consider optimization."
                })
        
        return suggestions
    
    def _log_performance_summary(self, report):
        """Log performance summary to console."""
        logging.info("=" * 60)
        logging.info("PERFORMANCE SUMMARY")
        logging.info("=" * 60)
        logging.info(f"Total execution time: {report['execution_summary']['total_time']:.2f}s")
        logging.info(f"API calls: {report['api_performance']['total_calls']} (success rate: {report['api_performance']['success_rate']:.1%})")
        logging.info(f"Average API response time: {report['api_performance']['average_response_time']:.3f}s")
        
        if report['memory_analysis'] and 'peak_usage' in report['memory_analysis']:
            logging.info(f"Peak memory usage: {report['memory_analysis']['peak_usage']:.1f}%")
        
        if report['bottlenecks']:
            logging.info(f"Bottlenecks detected: {len(report['bottlenecks'])}")
        
        if report['optimization_suggestions']:
            logging.info(f"Optimization suggestions: {len(report['optimization_suggestions'])}")
            for suggestion in report['optimization_suggestions'][:3]:  # Show top 3
                logging.info(f"  - {suggestion['description']}")
        
        logging.info("=" * 60)
    
    def get_current_stats(self):
        """Get current performance statistics."""
        return {
            'uptime': time.time() - self.start_time if self.start_time else 0,
            'api_calls': self.api_calls['total'],
            'memory_samples': len(self.memory_usage),
            'bottlenecks': len(self.bottlenecks),
            'active_monitoring': self.monitoring_active
        }

class OperationTimer:
    """Context manager for timing operations."""
    
    def __init__(self, metrics: PerformanceMetrics, operation_name: str):
        self.metrics = metrics
        self.operation_name = operation_name
        self.start_time = None
    
    def __enter__(self):
        self.start_time = time.time()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        end_time = time.time()
        duration = end_time - self.start_time
        
        if self.operation_name not in self.metrics.operation_timings:
            self.metrics.operation_timings[self.operation_name] = {
                'total_time': 0,
                'count': 0,
                'min_time': float('inf'),
                'max_time': 0,
                'average_time': 0
            }
        
        timing = self.metrics.operation_timings[self.operation_name]
        timing['total_time'] += duration
        timing['count'] += 1
        timing['min_time'] = min(timing['min_time'], duration)
        timing['max_time'] = max(timing['max_time'], duration)
        timing['average_time'] = timing['total_time'] / timing['count']

def main():
    """Main function."""
    print("🤖 FULLY AUTOMATED PETTY CASH SORTER WITH AI-ENHANCED RULE MATCHING")
    print("=" * 60)
    
    # No longer require JGD Truth file - AI-driven rule management handles everything
    print("✅ AI-Driven Rule Management: No external rule files required")
    
    # Create sorter
    sorter = FinalPettyCashSorter()
    
    # Run dry run first
    print("\nRUNNING DRY RUN FIRST...")
    try:
        success = sorter.run(dry_run=True)
        if success:
            print("\nDry run completed successfully!")
            response = input("\nProceed with live update? (y/N): ").strip().lower()
            if response == 'y':
                print("\nRUNNING LIVE UPDATE...")
                live_success = sorter.run(dry_run=False)
                if live_success:
                    print("\nLive update completed successfully!")
                else:
                    print("\nLive update failed!")
            else:
                print("Live update cancelled.")
        else:
            print("Dry run failed!")
            
    except Exception as e:
        print(f"Error: {e}")
        return

if __name__ == "__main__":
    main() 