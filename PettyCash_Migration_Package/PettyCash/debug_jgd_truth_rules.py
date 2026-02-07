#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Debug JGD Truth Rules Loading
"""

import openpyxl
import logging
from pathlib import Path

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def debug_jgd_truth_rules():
    """Debug the JGD Truth rules loading process."""
    
    jgd_truth_file = "JGD Truth Current.xlsx"
    
    if not Path(jgd_truth_file).exists():
        print(f"❌ JGD Truth file not found: {jgd_truth_file}")
        return
    
    print("🔍 DEBUGGING JGD TRUTH RULES LOADING")
    print("=" * 60)
    
    try:
        workbook = openpyxl.load_workbook(jgd_truth_file)
        print(f"✅ Loaded workbook with {len(workbook.sheetnames)} sheets: {workbook.sheetnames}")
        
        rules_cache = {}
        total_rules = 0
        
        for sheet_name in workbook.sheetnames:
            print(f"\n📋 Processing sheet: {sheet_name}")
            worksheet = workbook[sheet_name]
            
            print(f"   Sheet dimensions: {worksheet.max_row} rows x {worksheet.max_column} columns")
            
            # Check first few rows to understand structure
            print("   First 5 rows content:")
            for row in range(1, min(6, worksheet.max_row + 1)):
                row_data = []
                for col in range(1, min(6, worksheet.max_column + 1)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "None")
                print(f"     Row {row}: {row_data}")
            
            # Load rules (starting from row 2 as per the original logic)
            sheet_rules = 0
            for row in range(2, worksheet.max_row + 1):
                source = worksheet.cell(row=row, column=1).value
                target_sheet = worksheet.cell(row=row, column=2).value
                target_header = worksheet.cell(row=row, column=3).value
                
                if source and str(source).strip():
                    source_clean = str(source).strip()
                    rule_key = f"{source_clean}_{sheet_name}"
                    
                    rules_cache[rule_key] = {
                        'source': source_clean,
                        'target_sheet': str(target_sheet).strip() if target_sheet else '',
                        'target_header': str(target_header).strip() if target_header else '',
                        'jgd_sheet': sheet_name
                    }
                    
                    sheet_rules += 1
                    
                    # Show first few rules for each sheet
                    if sheet_rules <= 5:
                        print(f"     Rule {sheet_rules}: '{source_clean}' -> '{target_sheet}' / '{target_header}'")
            
            print(f"   ✅ Loaded {sheet_rules} rules from {sheet_name}")
            total_rules += sheet_rules
        
        workbook.close()
        
        print(f"\n📊 SUMMARY")
        print("=" * 60)
        print(f"Total rules loaded: {total_rules}")
        print(f"Rules cache size: {len(rules_cache)}")
        
        # Check for some specific rules that should exist
        print(f"\n🔍 CHECKING FOR SPECIFIC RULES")
        print("=" * 60)
        
        # Check for some of the missing sources from your CSV
        test_sources = [
            "SALE TO HIGH HOUSE",
            "PURE LABS", 
            "PAYROLL 26444",
            "ORDER SOONER WHOLESALE INC",
            "C-STORE",
            "WALMART",
            "VENMO"
        ]
        
        for test_source in test_sources:
            found = False
            for rule_key, rule in rules_cache.items():
                if rule['source'].upper() == test_source.upper():
                    print(f"✅ Found rule for '{test_source}': {rule['target_sheet']} / {rule['target_header']} (Company: {rule['jgd_sheet']})")
                    found = True
                    break
            
            if not found:
                print(f"❌ No rule found for '{test_source}'")
        
        # Show some sample rules by company
        print(f"\n📋 SAMPLE RULES BY COMPANY")
        print("=" * 60)
        
        companies = ['NUGZ', 'JGD', 'PUFFIN']
        for company in companies:
            company_rules = [rule for rule in rules_cache.values() if rule['jgd_sheet'] == company]
            print(f"\n{company} Rules ({len(company_rules)} total):")
            for i, rule in enumerate(company_rules[:10]):  # Show first 10
                print(f"  {i+1}. '{rule['source']}' -> '{rule['target_sheet']}' / '{rule['target_header']}'")
            if len(company_rules) > 10:
                print(f"  ... and {len(company_rules) - 10} more")
        
        return rules_cache
        
    except Exception as e:
        print(f"❌ Error debugging JGD Truth rules: {e}")
        return None

if __name__ == "__main__":
    debug_jgd_truth_rules() 